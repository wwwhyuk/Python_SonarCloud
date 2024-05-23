# -*- coding: utf-8 -*-
import string

import openpyxl
from openpyxl import load_workbook
import configparser
import concurrent.futures
import csv
import traceback
import inspect
import logging
import math
import shutil
import tkinter as tk
from collections import defaultdict
from tkinter import ttk, scrolledtext
from tkinter import filedialog
import os
from datetime import datetime, timedelta
from tkinter import messagebox
import time
import subprocess
import re  # Regular expression module
import threading  # For background tasks
import queue
from typing import Callable
from influxdb import InfluxDBClient
import pandas as pd
import sys
import mysql.connector
import xlsxwriter


class FileListApp:
    # init 실행시 초기화 함수
    def __init__(self, master):
        self.is_rotating = True
        self.log_folder = 'log'
        if not os.path.exists(self.log_folder):
            os.makedirs(self.log_folder)
        self.connection = None
        self.master = master
        self.master.title("천마 탐지추적장치 장비 데이터 업로드  V1.4.0 [2024-05-22 Released]")

        # 2024년 4월 24일
        # 선택 파일 제외 리스트
        self.select_files = []
        self.unselect_files = []
        self.new_filename = ''
        self.get_cpnt_id = ''
        self.fail_sig = 0
        self.connectiond = None
        self.connections = None
        self.stop_signal = 0

        # 2024년 5월 21일
        # 대상파일 선택 리스트
        self.selected_files = []
        self.select_signal = 1
        self.select_signal_one = 1

        # 2024년 5월 22일
        self.starting_time = 0
        self.message_fail_color = ''

        # 2024년 5월 23일
        self.file_select_signal = 1
        self.connect_ini_file_name = 'connect.ini'
        self.traceback_str = '\n=============== Traceback 내용 ===============\n'
        # var
        self.signal = 0  # 데이터 업로드인지, csv 변환인지 분기

        # Title label of treeview
        self.title_label = tk.Label(master, text="대상 파일 목록", font=('Helvetica', 12, 'bold'))
        self.title_label.grid(row=0, column=0, sticky='nw', pady=10, padx=10)

        # File list treeview
        self.file_tree = ttk.Treeview(master, columns=(
            'Type', 'Filename', 'Start Time', 'End Time', 'Count', 'Status', "Size", "Extension"),
                                      show='headings')
        self.file_tree.heading('Type', text='유형')
        self.file_tree.heading('Filename', text='파일명')
        self.file_tree.heading('Start Time', text='시작시간')
        self.file_tree.heading('End Time', text='완료시간')
        self.file_tree.heading('Count', text='처리건수')
        self.file_tree.heading('Status', text='상태')
        self.file_tree.heading('Size', text='크기')
        self.file_tree.heading('Extension', text='확장자')

        # Set the width of the 'Type' column
        self.file_tree.column('Type', width=50, anchor='center')
        self.file_tree.column('Filename', width=230)
        self.file_tree.column('Start Time', width=140, anchor='center')
        self.file_tree.column('End Time', width=140, anchor='center')
        self.file_tree.column('Count', width=80, anchor='center')
        self.file_tree.column('Status', width=100, anchor='center')
        self.file_tree.column('Size', width=0, stretch=False)  # Set width to 0
        self.file_tree.column('Extension', width=0, stretch=False)  # Set width to 0
        self.file_tree['height'] = 24

        # Create a vertical scrollbar
        vsb = ttk.Scrollbar(root, orient="vertical", command=self.file_tree.yview)
        self.file_tree.configure(yscrollcommand=vsb.set)
        vsb.grid(row=1, column=0, sticky="nes", padx=(0, 0))

        # Set the Treeview to use the scrollbars
        # vsb.config(command=self.file_tree.yview)

        self.file_tree.grid(row=1, column=0, pady=0, padx=0, sticky='new')

        # Log text widget
        # self.log_text = scrolledtext.ScrolledText(master, width=100, height=0)
        # self.log_text.grid(row=2, column=0, sticky='new', padx=(0, 0),
        #                    pady=(0, 0))
        # self.log_text.grid_remove()

        # Folder selection button
        self.select_folder_button = tk.Button(master, text="대상폴더 선택", command=self.select_folder, width=15, height=2)
        self.select_folder_button.grid(row=1, column=1, sticky='nw', pady=10, padx=10)

        # Mark Exclude button
        self.mark_exclude_button = tk.Button(master, text="대상파일 선택", command=self.mark_exclude, width=15, height=2,
                                             state='disabled')
        self.mark_exclude_button.grid(row=1, column=1, sticky='nw', pady=54, padx=10)

        # Mark Exclude button
        self.select_all_button = tk.Button(master, text="파일전체 선택", command=self.select_all, width=15, height=2,
                                           state='disabled')
        self.select_all_button.grid(row=1, column=1, sticky='nw', pady=98, padx=10)

        # Execute Upload button
        self.execute_upload_button = tk.Button(master, text="데이터 업로드", command=lambda: self.execute_upload(1), width=15,
                                               height=2,
                                               state='disabled')
        self.execute_upload_button.grid(row=1, column=1, sticky='nw', pady=142, padx=10)

        # Execute Upload button
        # self.execute_upload_button = tk.Button(master, text="데이터 업로드", command=self.extract_data, width=15,
        #                                        height=2,
        #                                        state='normal')
        # self.execute_upload_button.grid(row=1, column=1, sticky='nw', pady=98, padx=10)

        # Execute Upload button
        self.execute_translate_button = tk.Button(master, text="CSV 변환 / 결합", command=lambda: self.execute_upload(2),
                                                  width=15, height=2,
                                                  state='disabled')
        self.execute_translate_button.grid(row=1, column=1, sticky='nw', pady=186, padx=10)

        # Execute Upload button
        self.execute_insert_button = tk.Button(master, text="최근 7일치\n데이터 삽입", command=self.execute_insert,
                                               width=15, height=2,
                                               state='normal')
        self.execute_insert_button.grid(row=1, column=1, sticky='nw', pady=186, padx=10)
        self.execute_insert_button.grid_remove()

        # Close button
        self.close_button = tk.Button(master, text="닫기", command=self.close_window, width=15, height=2)
        self.close_button.grid(row=1, column=1, sticky='nw', pady=228, padx=10)

        # Progress bar
        self.progress_bar = ttk.Progressbar(master, orient="horizontal", length=100, mode="determinate")
        self.progress_bar.grid(row=2, column=0, sticky='ew', pady=10, padx=10)
        self.progress_bar.grid_remove()

        # Label for progressbar
        self.progressbar_label = tk.Label(master, text="0%", bd=0)
        self.progressbar_label.grid(row=2, column=0, pady=10, padx=10)
        self.progressbar_label.grid_remove()

        # Label for total processed files count
        self.total_files_label = tk.Label(master, text="Done/Total:")
        self.total_files_label.grid(row=2, column=1, sticky='w', pady=5, padx=10)
        self.total_files_label.grid_remove()

        # 로딩 인디케이터를 위한 캔버스 추가
        self.loading_canvas = tk.Canvas(master, width=50, height=50, highlightthickness=0)
        self.loading_canvas.grid(row=1, column=1, pady=(100, 0), padx=10)
        self.loading_arc = self.loading_canvas.create_arc(5, 5, 45, 45, start=0, extent=150, style='arc')
        self.loading_canvas.grid_remove()

        # 로딩 인디케이터를 위한 캔버스 추가
        self.loading_canvas1 = tk.Canvas(master, width=50, height=20, highlightthickness=0)
        self.loading_canvas1.grid(row=1, column=1, sticky='nwe', pady=(400, 0), padx=10)
        self.loading_bar = self.loading_canvas1.create_rectangle(0, 0, 0, 5, fill='blue', width=0)

        # 로그 설정 초기화
        self.setup_logger()

        self.path = ''

        # 캔버스 사이즈
        self.canvas_width = 100
        self.canvas_height = 75

        # 캔버스 추가
        self.animation_canvas = tk.Canvas(master, width=self.canvas_width, height=self.canvas_height)
        self.animation_canvas.grid(row=0, column=1, rowspan=3, padx=10, pady=(450, 0))

        # 초기 위치 설정 (캔버스 중앙)
        self.current_x = self.canvas_width // 2
        self.current_y = self.canvas_height // 2

        # 초기 반지름 설정
        self.radius = 50

        # 초기 각도 설정
        self.angle = 0

        # 표 생성
        self.file_tree2 = ttk.Treeview(master, height=5, columns='run', show='headings')
        self.file_tree2.heading('run', text='실행 중인 작업')
        self.file_tree2.column('run', width=100, anchor='center')
        self.file_tree2.grid(column=1, row=1, padx=10, pady=(280, 0), sticky='nwe')

        config = configparser.ConfigParser()
        config.read(self.connect_ini_file_name, 'UTF-8')
        get_type = config.get('Program', 'type')

        if get_type == 'ALL':
            print('ALL')
            self.equipment_data = self.mysql_load_equip_data()
            self.influx_connect = self.check_influxdb_connection()

        elif get_type == 'CSV':
            self.execute_upload_button.grid_remove()
            self.execute_translate_button.grid(row=1, column=1, sticky='nw', pady=142, padx=10)

    def select_all(self):
        if self.select_signal == 1:
            self.select_all_button.config(text="전체선택 취소")
            self.select_signal = 2
            all_items = self.file_tree.get_children()
            for item in all_items:
                print(item)
                file_info = self.file_tree.item(item, 'values')
                file_type, filename, start_time, end_time, count, status, size, extension = file_info
                status = "선택"
                self.file_tree.item(item, tags=('selected',),
                                    values=(file_type, filename, start_time, end_time, count, status, size, extension))
                self.file_tree.selection_add(item)
                self.file_tree.tag_configure('selected', background='orange')

        elif self.select_signal == 2:
            self.select_all_button.config(text="파일전체 선택")
            self.select_signal = 1
            all_items = self.file_tree.get_children()
            for item in all_items:
                print(item)
                file_info = self.file_tree.item(item, 'values')
                file_type, filename, start_time, end_time, count, status, size, extension = file_info
                status = ""
                self.file_tree.item(item, tags=('unselected',),
                                    values=(file_type, filename, start_time, end_time, count, status, size, extension))
                self.file_tree.selection_add(item)
                self.file_tree.tag_configure('unselected', background='white')

    def show_loading(self):
        self.progressing()
        self.animation_canvas.grid()

    def hide_loading(self):
        self.animation_canvas.grid_remove()

    # MySQL에서 장비 데이터 불러오는 함수
    def mysql_load_equip_data(self):
        print(' 결과 나와야하는데 ')
        """MySQL로부터 eqp_id, eqp_num을 로드하고 딕셔너리로 반환합니다."""
        try:
            # MySQL 연결 설정
            mysql_config = self.read_db_config(self.connect_ini_file_name, 'MySQL')

            # print(f"MySQL 연결 정보: {mysql_config['host']}//{mysql_config['port']}//{mysql_config['user']}//{mysql_config['password']}//{mysql_config['database']}")

            connection = mysql.connector.connect(
                host=mysql_config['host'],
                port=mysql_config['port'],
                user=mysql_config['user'],
                password=mysql_config['password'],
                database=mysql_config['database'],  # 사용할 데이터베이스 이름을 정확히 입력해주세요.
                connection_timeout=2  # 연결 타임아웃을 3초로 설정
            )
        except mysql.connector.Error as error:
            print(f"MySQL 연결 실패: {error}")
            logging.error(f"Failed to write data to InfluxDB: {error}")
            messagebox.showerror("MySQL 연결 실패", f"{error}")
            root.destroy()  # 윈도우를 닫아 프로그램을 종료합니다.
            err_msg = traceback.format_exc()
            print(err_msg)
            log_file_path = f'log/{datetime.now().strftime("%Y년%m월%d일%H시%M분%S초")}_MySqlError.log'
            with open(log_file_path, 'w') as log_file:
                log_file.write(str(error))
                log_file.write('self.traceback_str')
                log_file.write(err_msg)
            root.destroy()

            return {}  # 여기서 함수 종료
        try:
            cursor = connection.cursor()
            query = "SELECT eqp_id, eqp_num FROM tb_eqp_cl_cd"
            cursor.execute(query)

            # 결과를 딕셔너리로 변환
            equipment_data = {}
            for eqp_id, eqp_num in cursor:
                equipment_data[eqp_id] = eqp_num
            print("MySQL로부터 데이터를 로드하는데 성공했습니다:", equipment_data)

            cursor.close()
            connection.close()
            return equipment_data

        except mysql.connector.Error as error:
            print(f"MySQL로부터 데이터를 로드하는데 실패했습니다: {error}")
            logging.error(f"MySQL로부터 데이터를 로드하는데 실패했습니다: {error}")

            return {}

    # 로그 파일 생성 및 로깅 초기화 함수
    def setup_logger(self):
        """로그 설정을 초기화하고 로그 파일을 특정 폴더에 저장합니다."""
        # 'logs' 폴더 경로 확인 및 생성
        # log_directory = os.path.join(os.getcwd(), 'logs')
        # raw_log_directory = os.path.join(log_directory, 'raw')  # 'logs/raw' 폴더 경로
        # if not os.path.exists(log_directory):
        #     os.makedirs(log_directory)  # 'logs' 폴더가 없으면 생성
        # # if not os.path.exists(raw_log_directory):
        # #     os.makedirs(raw_log_directory)  # 'logs/raw' 폴더가 없으면 생성
        #
        # # 로그 파일 경로 생성
        # now = datetime.now()
        # log_filename = now.strftime('%Y%m%d'
        #                             # '_%H%M%S'
        #                             '.log')  # '20240222_140201.log' 형식
        # log_file_path = os.path.join(log_directory, log_filename)  # 로그 파일의 전체 경로

        # 로그 설정
        logging.basicConfig(level=logging.INFO,
                            format='%(asctime)s: %(message)s', datefmt='%Y-%m-%d %H:%M:%S')

        logging.info("업로드 시작")

    # 로깅 함수
    def log_file_upload(self, filename, start_time, end_time, count, status, size, extension):
        """각 파일 업로드 정보를 로그 파일에 기록합니다."""
        # logging.info(f"파일명: {filename}, 시작시간: {start_time}, 완료시간: {end_time}, 처리건수: {count}, 상태: {status}, 파일크기: {size}KB, 확장자: {extension}")

    # 로깅 마무리 함수
    def finalize_logging(self, total_files):
        """업로드 요약 정보를 로그 파일에 기록합니다."""
        logging.info(f"업로드 완료. 총 파일 개수: {total_files} \n")

    # 폴더 선택 버튼 클릭 이벤트 함수
    def select_folder(self):
        self.select_files = []
        self.unselect_files = []
        self.selected_files = []
        self.new_filename = ''
        self.get_cpnt_id = ''
        self.fail_sig = 0
        self.connectiond = None
        # 인플럭스DB 연결 확인
        # if not self.check_influxdb_connection():
        #     messagebox.showerror("Influx DB Connection Error", "\"connnect.ini\" 파일의 DB정보와 Path를 확인해주세요.")
        #     return  # DB 연결 실패시 함수 종료
        # Open folder selection dialog

        self.select_all_button.config(text="파일전체 선택")
        self.select_signal = 1

        self.folder_path = filedialog.askdirectory(title="대상폴더 선택")
        if self.folder_path:
            # 특수문자 검사 로직 추가
            if any(char in self.folder_path for char in ['%', '&', '*', '$', '@', '!']):
                messagebox.showerror("Invalid Folder Path", "폴더 경로에 특수문자가 포함되어 있습니다.")
                # 특수문자가 포함된 경우 여기서 함수 종료
                return
            # Load file list for the selected folder
            self.load_file_list(self.folder_path)
            self.mark_exclude_button['state'] = 'normal'
            self.select_all_button['state'] = 'normal'
            self.execute_upload_button['state'] = 'normal'
            self.execute_translate_button['state'] = 'normal'

            directory_path = self.folder_path
            for filename in os.listdir(directory_path):
                if filename.endswith(".csv") and "_IU_" in filename:
                    file_path = os.path.join(directory_path, filename)
                    os.remove(file_path)
                if filename.endswith(".txt"):
                    file_path = os.path.join(directory_path, filename)
                    os.remove(file_path)
            print('삭제 완료')

        self.is_rotating = False

    # 선택된 파일리스트 테이블에 로딩 함수
    def load_file_list(self, folder_path):
        # Add file list to the treeview
        self.file_tree.delete(*self.file_tree.get_children())  # Clear the existing list
        self.path = self.folder_path

        # Initialize ProgressBar
        self.progress_bar["value"] = 0
        self.progress_bar["maximum"] = 0

        self.progressbar_label.config(text="0%")

        files = [f for f in os.listdir(folder_path)
                 if f.endswith('.raw') or (f.endswith('.csv') and "_IU_" not in f)]
        for file in files:
            file_path = os.path.join(folder_path, file)
            file_size_bytes = os.path.getsize(file_path)  # 파일 크기(바이트)
            file_size_kb = file_size_bytes / 1024  # 파일 크기를 KB 단위로 변환
            file_size_kb_rounded = round(file_size_kb, 1)  # 소수점 첫 번째 자리까지 반올림
            file_extension = os.path.splitext(file)[1]  # 파일 확장자
            # Update Type based on filename
            if "_IU_" in file:
                file_type = '상태'
            elif "_MI_" in file:
                file_type = '계측'
            else:
                file_type = '기타'

            file_start_time = ""
            file_end_time = ""
            file_count = ""
            file_status = ""

            self.file_tree.insert('', 'end',
                                  values=(file_type, file, file_start_time, file_end_time, file_count, file_status,
                                          file_size_kb_rounded, file_extension))

    # 파일 제외 처리 함수
    def mark_exclude(self):

        """ 2024년 5월 23일 '이미 선택된 상황에 대해서 처리' """
        # selected_items = self.file_tree.selection()
        # for item in selected_items:
        #     item_info = self.file_tree.item(item)
        #     values = item_info['values']
        #     status = values[5] if len(values) > 5 else 'No status available'
        #     print(status)
        # if status is "선택":
        #     self.mark_exclude_button.config(text="대상파일 취소")
        #     print('not have')
        # else:
        #     print('what')


        selected_items = self.file_tree.selection()
        if selected_items:
            print('뭔데', selected_items)
            for item in selected_items:
                file_info = self.file_tree.item(item, 'values')
                file_type, filename, start_time, end_time, count, status, size, extension = file_info

                status = "선택"
                self.file_tree.item(item, tags=('marked',),
                                    values=(file_type, filename, start_time, end_time, count, status, size, extension))

                # 'marked' 태그가 설정된 아이템에 대해 오렌지색 배경 설정
                self.file_tree.tag_configure('marked', background='orange')
        else:
            messagebox.showinfo('선택된 파일 없음', '선택된 파일이 없습니다.\n파일을 선택해주십시요.')

    # 로딩인디케이터 에니매이션 처리 함수
    def animate_loading(self, step):
        self.loading_canvas.itemconfig(self.loading_arc, start=step, extent=150)
        self.master.after(100, self.animate_loading, (step + 10) % 360)

    # 인플럭스db 연결성 및 path 체크 함수(폴더선택시 체크)
    def check_influxdb_connection(self):
        try:
            # 인플럭스DB 설정 읽기
            db_config = self.read_db_config()

            # influx.exe 파일 경로 확인
            influx_exe_path = os.path.join(db_config['path'], 'influx.exe')
            if not os.path.exists(influx_exe_path):
                raise ValueError(f"influx.exe not found in path: {db_config['path']}")

            # 인플럭스DB 클라이언트 설정
            client = InfluxDBClient(host=db_config['host'],
                                    port=db_config['port'],
                                    database=db_config['database'])

            client.close()
            return True
        except Exception as e:
            err_msg = traceback.format_exc()
            print(err_msg)
            log_file_path = f'log/{datetime.now().strftime("%Y년%m월%d일%H시%M분%S초")}_InfluxdbError.log'
            with open(log_file_path, 'w') as log_file:
                log_file.write(str(e))
                log_file.write('self.traceback_str')
                log_file.write(err_msg)
            print(f"DB Connection Error: {e}")
            messagebox.showerror("InfluxDB 연결 실패", f"{e}")
            root.destroy()
            return False

    # 업로드 버튼 클릭 이벤트 함수
    def execute_upload(self, sig):
        self.message_fail_color = ''
        self.starting_time = 0

        self.starting_time = time.time()

        total_files = sum(
            1 for item in self.file_tree.get_children() if "선택" in self.file_tree.item(item, 'values')[5])

        # "선택"이라는 데이터가 있는지 확인
        selection_exists = any("선택" in self.file_tree.item(item, 'values')[5] for item in self.file_tree.get_children())

        # 만약 "선택"이라는 데이터가 없으면 "없음" 출력
        if not selection_exists:
            messagebox.showinfo('선택된 파일 없음', '선택된 파일이 없습니다.\n대상 파일을 선택해주십시요.')
            return

        self.stop_signal = 0
        threading.Thread(target=self.progressing).start()
        self.animation_canvas.grid()
        self.signal = sig
        if sig == 1:
            self.file_tree2.insert('', 'end', iid="1", values=('데이터 업로드',))
        elif sig == 2:
            self.file_tree2.insert('', 'end', iid="2", values=('CSV 변환 / 결합',))

        self.processed_files_count = 0
        self.stime = time.time()
        self.upload_queue = queue.Queue()
        for i, item in enumerate(self.file_tree.get_children()):
            self.upload_queue.put((i, item))

        # 진행 상태 추적 및 UI 업데이트
        self.track_progress(total_files)
        self.execute_translate_button['state'] = 'disabled'
        self.execute_upload_button['state'] = 'disabled'
        self.mark_exclude_button['state'] = 'disabled'
        self.select_all_button['state'] = 'disabled'
        self.select_folder_button['state'] = 'disabled'

        # 스레드 시작
        for _ in range(total_files):
            worker_thread = threading.Thread(target=self.upload_worker, args=(sig,))
            worker_thread.daemon = True
            worker_thread.start()

    # 작업큐에 저장된 upload_file 함수를 순차적으로 호출하는 함수
    def upload_worker(self, sig):
        while not self.upload_queue.empty():
            i, item = self.upload_queue.get()
            try:
                self.upload_file(item, sig)
            finally:
                self.upload_queue.task_done()

    # 파일 업로드 로직을 수행 하는 함수
    def upload_file(self, item, sig):
        try:

            file_info = self.file_tree.item(item, 'values')
            file_type, filename, start_time, end_time, count, status, size, extension = file_info

            if "선택" in status:
                start_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                self.file_tree.item(item,
                                    values=(file_type, filename, start_time, end_time, count, status, size, extension))
                if 'IU' in filename:
                    self.new_filename = filename.replace('.raw', '.csv')
                # self.unselect_files.append(self.new_filename)
                self.selected_files.append(self.new_filename)
                if sig == 1:
                    if filename.endswith('.raw'):
                        # 파일명에서 eqp_num 추출
                        eqp_num_from_filename = filename[7:10]
                        if self.get_eqp_id_by_eqp_num(eqp_num_from_filename):
                            eqp_id = self.get_eqp_id_by_eqp_num(eqp_num_from_filename)
                        else:
                            eqp_id = eqp_num_from_filename

                        # 실행 파일이 위치한 경로를 찾습니다.
                        if getattr(sys, 'frozen', False):
                            # PyInstaller로 생성된 경우
                            application_path = os.path.dirname(sys.executable)
                        else:
                            # 스크립트를 직접 실행하는 경우 (예: 파이참에서)
                            application_path = os.path.dirname(__file__)
                        # raw 아니라면 아래 sig 분기는 상관없음
                        db_config = self.read_db_config()
                        old_file_path = os.path.join(self.folder_path, filename)
                        new_file_path = old_file_path.replace('.raw', '.txt')
                        upload_exe_path = os.path.join(application_path, 'CBM_IUData_Converter(RtoT)_1.2.exe')
                        try:
                            # 20240326 C++ 프로그램 호출시에도 host, port 파라미터 추가 전달
                            process = subprocess.run(
                                [upload_exe_path, old_file_path, new_file_path, db_config['database'],
                                 db_config['path'],
                                 eqp_id, db_config['host'], db_config['port']], capture_output=True, text=True,
                                creationflags=subprocess.CREATE_NO_WINDOW)
                            logging.info(process)

                            connectdb = (
                                f"C: && cd {db_config['path']} && influx -execute \"CREATE DATABASE {db_config['database']}\""
                                f" -host={db_config['host']} -port={db_config['port']}")

                            command = (
                                f"C: && cd {db_config['path']} && influx -import -path=\"{new_file_path}\" -database={db_config['database']}"
                                f" -host={db_config['host']} -port={db_config['port']}")

                            subprocess.run(connectdb, shell=True)
                            subprocess.run(command, shell=True)
                        except subprocess.CalledProcessError as e:
                            # 실행 실패 시 에러 로깅
                            print(f"Error executing uploadTxt.exe: {e.output}")

                        # process.communicate()
                        with open(new_file_path, 'r') as f:
                            count_all = sum(1 for _ in f)
                            # 4에서 3으로 변경
                            count = count_all - 3

                            if count_all == 4:
                                count = 0

                            # print(f"new_file_path: {new_file_path}, countAll: {countAll}, count: {count}")

                            # 20240312 데이터가 없는 건들은 마이너스 값이 나와서 0으로 처리
                            if count < 0:
                                count = 0

                    elif filename.endswith('.csv'):
                        csv_file_path = os.path.join(self.folder_path, filename)
                        self.upload_csv_to_influxdb(csv_file_path, filename)

                        with open(csv_file_path, 'r') as f:
                            count = sum(1 for _ in f) - 1

                elif sig == 2:
                    self.get_cpnt_id = filename[0:10]
                    print('여기는 장비명입니다.', self.get_cpnt_id)
                    if filename.endswith('.raw'):
                        # 파일명에서 eqp_num 추출
                        eqp_num_from_filename = filename[7:10]
                        print('여기를 읽으면 안돼')
                        eqp_id = eqp_num_from_filename

                        if getattr(sys, 'frozen', False):
                            # PyInstaller로 생성된 경우
                            application_path = os.path.dirname(sys.executable)
                        else:
                            # 스크립트를 직접 실행하는 경우 (예: 파이참에서)
                            application_path = os.path.dirname(__file__)
                        # raw 아니라면 아래 sig 분기는 상관없음

                        db_config = {
                            'database': " ",
                            'path': " ",
                            'host': " ",
                            'port': " "
                        }

                        old_file_path = os.path.join(self.folder_path, filename)
                        new_file_path = old_file_path.replace('.raw', '.csv')
                        upload_exe_path = os.path.join(application_path, 'CBM_IUData_Converter(RToC)_1.1.exe')
                        try:
                            # 20240326 C++ 프로그램 호출시에도 host, port 파라미터 추가 전달
                            process = subprocess.run(
                                [upload_exe_path, old_file_path, new_file_path, db_config['database'],
                                 db_config['path'],
                                 eqp_id, db_config['host'], db_config['port']], capture_output=True, text=True,
                                creationflags=subprocess.CREATE_NO_WINDOW)
                            logging.info(process)
                        except subprocess.CalledProcessError as e:
                            # 실행 실패 시 에러 로깅
                            print(f"Error executing uploadTxt.exe: {e.output}")

                        # process.communicate()
                        with open(new_file_path, 'r') as f:
                            count_all = sum(1 for _ in f)
                            # 4에서 3으로 변경
                            count = count_all - 3

                            if count_all == 4:
                                count = 0

                            # print(f"new_file_path: {new_file_path}, countAll: {countAll}, count: {count}")

                            # 20240312 데이터가 없는 건들은 마이너스 값이 나와서 0으로 처리
                            if count < 0:
                                count = 0

                    elif filename.endswith('.csv'):
                        print()
                        csv_file_path = os.path.join(self.folder_path, filename)
                        if 'MI' in filename:
                            # self.unselect_files.append(filename)
                            self.selected_files.append(filename)

                        # self.upload_csv_to_influxdb(csv_file_path, filename)
                        #
                        with open(csv_file_path, 'r') as f:
                            count = sum(1 for _ in f) - 1

                end_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                status = 'Done'
                self.file_tree.item(item,
                                    values=(file_type, filename, start_time, end_time, count, status, size, extension))
                self.processed_files_count += 1
                file_info = self.file_tree.item(item, 'values')
                print(file_info, "uploaded files")
                # 로그에 파일 정보 기록
                self.log_file_upload(filename, start_time, end_time, count, status, size, extension)

        except Exception as E:
            print('error : ', E)
            err_msg = traceback.format_exc()
            print(err_msg)
            log_file_path = f'log/{datetime.now().strftime("%Y년%m월%d일%H시%M분%S초")}_TranslateError.log'
            with open(log_file_path, 'w') as log_file:
                log_file.write(str(E))
                log_file.write('self.traceback_str')
                log_file.write(err_msg)
            messagebox.showinfo("오류", f"아래와 같은 오류가 발생하였습니다. log 폴더 내 log 파일을 확인해주십시요.\n{E}")
            messagebox.showinfo("오류", "가급적 종료하여 주시고, 오류 해결 후 실행해주십시요.")
            self.close_button['state'] = 'normal'

    # 업로드 결과를 MySQL에 저장 함수
    def mysql_upload(self):

        try:
            mysql_config = self.read_db_config(self.connect_ini_file_name, 'MySQL')

            connection = mysql.connector.connect(
                host=mysql_config['host'],
                port=mysql_config['port'],
                user=mysql_config['user'],
                password=mysql_config['password'],
                database=mysql_config['database'],  # 사용할 데이터베이스 이름을 정확히 입력해주세요.
                connection_timeout=2  # 연결 타임아웃을 3초로 설정
            )
        except mysql.connector.Error as error:
            print(f"MySQL 연결 실패: {error}")
            return  # 여기서 함수 종료

        try:
            cursor = connection.cursor()
            # self.stime이 유닉스 타임스탬프를 나타내는 float 또는 int 값일 때
            stime_datetime = datetime.fromtimestamp(self.stime)

            # MySQL에 삽입하기 위해 문자열 형식으로 변환 (예: 'YYYY-MM-DD HH:MM:SS')
            stime_str = stime_datetime.strftime('%Y-%m-%d %H:%M:%S')
            now_time = datetime.now()
            now_time_round = now_time.replace(microsecond=0)

            total_files = sum(
                1 for item in self.file_tree.get_children() if "선택" in self.file_tree.item(item, 'values')[5])

            # 업로드가 완료된 파일 정보를 tb_file_upload에 업로드
            sql = "INSERT INTO tb_file_upload (up_date, st_time, en_time, file_cnt, trsf_status, trsf_tp) VALUES (%s, %s, %s, %s, %s, %s)"
            cursor.execute(sql, (datetime.now().date(), stime_str, now_time_round, total_files, "Done", "UL"))
            print(self.stime, "start_time", stime_str)
            up_seq = cursor.lastrowid
            connection.commit()

            for index, item in enumerate(self.file_tree.get_children(), start=1):
                file_info = self.file_tree.item(item, 'values')
                file_type, filename, start_time, end_time, count, status, size, _ = file_info

                if status == "선택":
                    continue

                if file_type == "계측":
                    file_type = "MI"
                elif file_type == "상태":
                    file_type = "IU"
                else:
                    file_type = "Unknown"

                # 파일 건수가 빈 문자열인 경우 0으로 대체
                if count == '':
                    count = 0

                # 파일명에서 eqp_num 추출
                eqp_num_from_filename = filename[7:10]
                if self.get_eqp_id_by_eqp_num(eqp_num_from_filename):
                    eqp_id = self.get_eqp_id_by_eqp_num(eqp_num_from_filename)
                else:
                    eqp_id = ""

                # 업로드된 파일의 상세 정보를 tb_file_upload_det에 업로드
                sql = (
                    "INSERT INTO tb_file_upload_det (up_seq,up_det_seq, st_time, en_time, file_name, data_cnt, success_cnt, fail_cnt, trsf_status ,file_size, size_unit, data_type, eqp_id) "
                    "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)")
                cursor.execute(sql,
                               (up_seq, index, start_time, end_time, filename, int(count), int(count), 0, status,
                                size, "KB", file_type, eqp_id))
                print(file_info, "maria DB")
                connection.commit()

        except mysql.connector.Error as error:
            print(f"MySQL에 정보를 업로드하는 동안 오류가 발생했습니다: {error}")

        finally:
            if connection.is_connected():
                cursor.close()
                connection.close()

    # csv file upload 함수 thread 생성 후 호출 함수
    def upload_csv_to_influxdb(self, csv_file_path, filename):

        def upload():
            self.insert_csv_to_influxdb(csv_file_path, filename)

        thread = threading.Thread(target=upload)
        thread.start()

    # csv file upload 함수
    def insert_csv_to_influxdb(self, csv_file_path, filename):
        client = self.get_influxdb_client()

        measurement = "mi_data"  # 적절한 측정값 이름 설정

        # 파일명에서 eqp_num 추출
        eqp_num_from_filename = filename[7:10]
        if self.get_eqp_id_by_eqp_num(eqp_num_from_filename):
            eqp_id = self.get_eqp_id_by_eqp_num(eqp_num_from_filename)
        else:
            eqp_id = eqp_num_from_filename

        print(eqp_num_from_filename, eqp_id, "eqp_num", "eqp_id")
        df = pd.read_csv(csv_file_path)

        # 데이터프레임에 eqp_id 컬럼이 이미 존재하는지 확인하고, 없으면 추가
        if 'eqp_id' not in df.columns:
            df['eqp_id'] = eqp_id
        else:
            df['eqp_id'] = df['eqp_id'].fillna(eqp_id)  # 존재하는 경우, NaN 값만 채웁니다.

        for _, row in df.iterrows():
            timestamp = self.parse_timestamp(row['Time'])
            if timestamp is None:
                continue

            # eqp_id 컬럼 제외하고 나머지 값들을 float으로 변환하여 fields 딕셔너리에 추가
            fields = {key: float(value) for key, value in row.items() if
                      key != 'Time' and pd.notna(value) and key != 'eqp_id'}
            fields['eqp_id'] = str(row['eqp_id'])  # eqp_id를 fields에 추가

            data = [{
                "measurement": measurement,
                "fields": fields,
                "time": timestamp
            }]

            try:
                client.write_points(data)
            except Exception as e:
                logging.error(f"Failed to write data to InfluxDB: {e}")

        client.close()
        mess = ("InfluxUploadSuccess" + str(client))
        logging.info(mess)

    # eqp_numd으로 eqp_id 조회하는 함수
    def get_eqp_id_by_eqp_num(self, eqp_num):
        """eqp_num 값으로 eqp_id를 조회하여 반환합니다."""
        # equipment_data 딕셔너리에서 eqp_num에 해당하는 eqp_id를 찾습니다.
        for id, num in self.equipment_data.items():
            if num == eqp_num:
                return id  # 일치하는 eqp_id 반환
        return None  # 일치하는 항목이 없으면 None 반환

    # ini 파일에서 db커넥션 정보 가져오는 함수
    @staticmethod
    def read_db_config(filename, section):
        parser = configparser.ConfigParser()
        parser.read(filename, 'UTF-8')
        db = {}
        if parser.has_section(section):
            items = parser.items(section)
            for item in items:
                db[item[0]] = item[1]
        else:
            raise Exception(f'{section} not found in the {filename} file')

        if section == 'InfluxDB':
            db["database"] = "myTest"
        return db

    # InfluxDB 클라이언트 설정 함수
    def get_influxdb_client(self):
        db_config = FileListApp.read_db_config()  # Static method call

        # InfluxDB 클라이언트 생성
        client = InfluxDBClient(host=db_config['host'],
                                port=db_config['port'],
                                database=db_config['database'])
        return client

    # 타임스탬프 변환 함수
    @staticmethod
    def parse_timestamp(timestamp_str):
        if isinstance(timestamp_str, str) and re.match(r'\d{8}_\d{9}', timestamp_str):

            date_part, time_part = timestamp_str.split('_')

            year = int(date_part[:4])
            month = int(date_part[4:6])
            day = int(date_part[6:8])

            hour = int(time_part[:2])
            minute = int(time_part[2:4])
            second = int(time_part[4:6])
            millisecond = int(time_part[6:9])
            microsecond = millisecond * 1000

            dt = datetime(year, month, day, hour, minute, second, microsecond)

            return dt.isoformat()
        else:
            return None

    # 진행상황 UI(프로그레스바 등) 업데이트 함수
    def track_progress(self, total_files):
        def update_ui():
            progress_value = int((self.processed_files_count / total_files) * 100)
            self.progress_bar["value"] = progress_value
            if self.processed_files_count < total_files:
                self.master.after(100, update_ui)
            else:
                try:

                    if self.signal == 1:
                        self.mysql_upload()
                        self.extract_data()
                    elif self.signal == 2:
                        self.open_csv()
                        print('hello')
                except Exception as E:
                    print('error : ', E)
                    err_msg = traceback.format_exc()
                    print(err_msg)
                    log_file_path = f'log/{datetime.now().strftime("%Y년%m월%d일%H시%M분%S초")}_ProgressError.log'
                    with open(log_file_path, 'w') as log_file:
                        log_file.write(str(E))
                        log_file.write('self.traceback_str')
                        log_file.write(err_msg)
                    messagebox.showinfo("오류", f"아래와 같은 오류가 발생하였습니다. log 폴더 내 log 파일을 확인해주십시요.\n{E}")
                    messagebox.showinfo("오류", "가급적 종료하여 주시고, 오류 해결 후 실행해주십시요.")
                    self.close_button['state'] = 'normal'

        update_ui()

        # 업로드 요약 정보 로그 기록
        self.finalize_logging(total_files)

    def estimate_time(self):

        txt_files = [file for file in os.listdir(self.path) if file.endswith('.raw')]
        self.time = len(txt_files)
        if self.time < 90:
            self.time = '15분미만'
        else:
            self.time = '15분이상'

    def animate_loading1(self):
        print("test")

    def progressing(self):
        if self.stop_signal == 1:
            return
        # 이전에 그린 원 삭제
        self.animation_canvas.delete("rotating_circle")

        # 각도 증가
        self.angle += 4

        # 원 그리기
        x0 = self.current_x + self.radius * math.cos(math.radians(self.angle))
        y0 = self.current_y + self.radius * math.sin(math.radians(self.angle))
        x1 = self.current_x - self.radius * math.cos(math.radians(self.angle))
        y1 = self.current_y - self.radius * math.sin(math.radians(self.angle))
        self.animation_canvas.create_line(x0, y0, x1, y1, fill="orange", width=5, tags="rotating_circle")

        # 재귀적으로 자기 자신을 호출하여 반복 실행
        self.master.after(1, self.progressing)

    def extract_data(self):
        thread1 = threading.Thread(target=self.progressing)
        thread1.start()
        thread = threading.Thread(target=self.data_extraction_task)
        thread.start()

    def data_extraction_task(self):
        self.run_sub(self.path)
        self.hide_loading()
        self.file_tree2.delete("1")
        if self.fail_sig == 1:
            self.stop_signal = 1
            messagebox.showerror("실패", "데이터 업로드 실패!")
        elif self.fail_sig == 0:
            self.stop_signal = 1
            messagebox.showinfo("성공", f"데이터 업로드 완료!\n소요 시간 : {round(time.time() - self.starting_time)}초")
        self.select_folder_button['state'] = 'normal'
        self.execute_upload_button['state'] = 'disabled'

    def run_sub(self, path):
        try:
            self.main(path)
            value = path
            input_directory = os.path.join(value, 'output_sr_error.txt')
            output_directory = os.path.join(value, 'output_sr_result.txt')
            self.working_query(value, input_directory, output_directory)

            input_directory = os.path.join(value, 'output_tr_error.txt')
            output_directory = os.path.join(value, 'output_tr_result.txt')
            self.working_query(value, input_directory, output_directory)

            config = configparser.ConfigParser()
            config.read(self.connect_ini_file_name, 'UTF-8')
            get_type = config.get('Program', 'CHECK')

            merged = os.path.join(value, 'output_merge.txt')
            arrayed = os.path.join(value, 'output_array.txt')
            date_data = os.path.join(value, 'data by date')
            works = os.path.join(value, 'output_works.txt')
            result = os.path.join(value, 'output_result.txt')
            sr_error = os.path.join(value, 'output_sr_error.txt')
            tr_error = os.path.join(value, 'output_tr_error.txt')
            sr_result = os.path.join(value, 'output_sr_result.txt')
            tr_result = os.path.join(value, 'output_tr_result.txt')
            working = os.path.join(value, 'output_working.txt')

            files_to_move = [
                sr_error,
                tr_error,
                tr_result,
                sr_result,
                working
            ]

            files_to_delete_ori = [
                merged,
                arrayed,
                date_data,
                works,
                result,
                sr_error,
                tr_error,
                tr_result,
                sr_result,
                working
            ]

            files_to_delete_sub = [
                merged,
                arrayed,
                date_data,
                works,
                result,
                sr_error,
                tr_error,
                tr_result,
                sr_result,
                working
            ]

            if get_type == 'YES':
                result_directory = '.\\추출 결과\\'
                os.makedirs(result_directory, exist_ok=True)

                """ 2024년 5월 23일 파일 설명 생성 관련 """
                # # 파일명과 설명을 지정
                # file_name = '파일 설명.txt'
                #
                # # 파일 경로 설정
                # file_path = os.path.join(result_directory, file_name)
                #
                # # 파일이 존재하지 않는다면 새로 생성하고 'hello world' 추가
                # if not os.path.exists(file_path):
                #     with open(file_path, 'w', encoding='UTF-8') as file:
                #         file.write('\t✅FILE DESCRIPTION✅\t\n'
                #                    '\t 2024년4월26일 작성\n\n')
                #         file.write('▷ sr_error.txt : SR ( 탐지 레이더 ) 의 모든 상태 정보\n')
                #         file.write('▷ sr_result.txt : SR ( 탐지 레이더 ) 의 오류 상태 정보\n\n')
                #         file.write('▷ tr_error.txt : TR ( 추적 레이더 ) 의 모든 상태 정보\n')
                #         file.write('▷ tr_result.txt : TR ( 추적 레이더 ) 의 오류 상태 정보\n\n')
                #         file.write('▷ output_summary.txt : SR, TR, RCS 데이터의 날짜별 운용 시작, 종료 정보\n')
                #         file.write('▷ working_history.txt : SR, TR, RCS 데이터의 날짜별 운용 시작, 종료 정보(for QUERY_INSERT)\n\n')
                #         file.write('▶ 수집된 데이터가 존재하지않는 경우에는 결과 파일의 데이터가 존재하지 않습니다.')

                timestamped_directory = self.create_timestamped_directory(result_directory)

                self.move_files_to_directory(files_to_move, timestamped_directory)

                self.delete_directory(files_to_delete_sub)
            else:
                self.delete_directory(files_to_delete_ori)

            self.flag = False
        except Exception as E:
            self.flag = False
            self.fail_flag = 1
            err_msg = traceback.format_exc()
            print(err_msg)
            log_file_path = f'log/{datetime.now().strftime("%Y년%m월%d일%H시%M분%S초")}ProgressError.log'
            with open(log_file_path, 'w') as log_file:
                log_file.write(str(E))
                log_file.write('self.traceback_str')
                log_file.write(err_msg)
            self.close_button['state'] = 'normal'

    def read_csv(self, directory):
        for filename in self.selected_files:
            print(' 실행되어야 합니다. : \t', filename)
            if "_IU_" in filename and filename.endswith(".csv"):
                file_path = os.path.join(directory, filename)
                self.modify_csv(file_path)

    def modify_csv(self, file_path):
        temp_file = file_path + '.temp'

        csv_file = 'description_ini_csv.csv'

        # 1인 경우의 개수를 저장할 변수
        count_1 = 0
        # CSV 파일의 열의 개수를 저장할 변수
        num_columns = 0

        # CSV 파일 열기
        with open(csv_file, newline='') as file:
            reader = csv.reader(file)
            for row in reader:
                # 첫 번째 행의 열의 개수를 가져옴
                num_columns = len(row)
                # 첫 번째 열의 값이 1인 경우 count_1을 증가시킴
                if row[0] == '1':
                    count_1 += 1
        print("1인 경우의 개수:", count_1)
        print("CSV 파일의 열의 개수:", num_columns)

        with open(temp_file, 'w', encoding='utf-8') as temp:
            with open(file_path, 'r', encoding='utf-8') as file:
                for row in file:
                    comma_count = row.count(',')
                    if comma_count >= count_1 + 1:
                        corrected_row = row[::-1].replace(',', '', comma_count - count_1)[::-1]
                        temp.write(corrected_row)
                    else:
                        temp.write(row)
        shutil.move(temp_file, file_path)

    def merge_csv(self, file_path):
        try:
            # fixed_list 고정 컬럼 ( 헤더 )
            new_data = [
                "Description",
                "TR_Cabinet_temperatureFail=1",
                "TRadar_Fail=1",
                "SR_TX/RX_Fail=1",
                "SR_TX/RX_Fail=1",
                "SR_Antenna_Fail=1",
                "SR_Interface_Fail=1",
                "SR_Booster_Fail=1",
                "SR_Module_Booster_Fail=1",
                "Sradar_Fail=1",
                "TT_2nd_EL_stop_Fail=1",
                "TT_1st_EL_stop_Fail=1",
                "TT_1st_EL_High_stop_Fail=1",
                "SCD_Mask_Fail=1",
                "SCD_Elev_Fail=1",
                "SCD_AZ_Fail=1",
                "SCD_Fail=1"
            ]

            # 파일 경로 내 csv 파일들을 읽어들이기 위한 리스트 초기화
            list_of_file_names = []

            # self.unselect_files 리스트에 있는 각 파일명에 대해 파일 경로와 결합하여 리스트에 추가
            # for filename in self.unselect_files:
            for filename in self.selected_files:
                print(' 합산되어야 합니다. \t', filename)
                full_file_path = os.path.join(file_path, filename)
                if os.path.isfile(full_file_path) and filename.endswith(".csv"):
                    list_of_file_names.append(full_file_path)

            # 현재 날짜 가져오기
            current_date = datetime.now().strftime("%Y년%m월%d일%H시%M분%S초")

            # "_IU_"이 들어간 파일들을 합치기
            iu_dfs = []
            for file_name in list_of_file_names:
                if "_IU_" in file_name:
                    df = pd.read_csv(file_name)
                    iu_dfs.append(df)

            # "_MI_"이 들어간 파일들을 합치기
            mi_dfs = []
            for file_name in list_of_file_names:
                if "_MI_" in file_name:
                    df = pd.read_csv(file_name)
                    mi_dfs.append(df)

            # 빈 DataFrame인 경우 pass
            iu_dfs = [df for df in iu_dfs if df is not None]
            mi_dfs = [df for df in mi_dfs if df is not None]

            # OneDrive 때문에 오류 있음
            # desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
            # desktop_path = "C:\\Users\\rkfpt\\바탕 화면"
            config = configparser.ConfigParser()
            config.read(self.connect_ini_file_name, 'UTF-8')
            desktop_path = config.get('ExcelPath', 'path')
            fixed_check = config.get('Program', 'fixed_list')
            cell_check = config.get('Program', 'cell_condition')
            path_list = []

            # 데이터 프레임 병합
            if iu_dfs:
                xlsx_files = []
                iu_result_df = pd.concat(iu_dfs)
                iudf1 = iu_result_df
                iudf1['TIME'] = pd.to_datetime(iudf1['TIME'], format='%y%m%d_%H%M%S%f')
                iudf1 = iudf1.sort_values(by='TIME')
                iudf1['TIME'] = iudf1['TIME'].dt.strftime('%y%m%d %H%M%S.%f')
                result_path = f"{desktop_path}\\{self.get_cpnt_id}_IU_MERGED_{current_date}.csv"
                iudf1.to_csv(result_path, index=False)
                path_list.append(result_path)

                csv_file = result_path

                """ 2024년 5월 22일 CSV 서식 지정 시작 """
                """ 1) 서식 지정 YES 일때, 로직 실행 (CSV and XLSX 산출)
                    2) 서식 지정 NO 일때, 로직 실행 X (CSV만 산출) -> pass
                    3) 서식 지정이 YES|NO 일때 처리가 끝난 뒤,
                        fixed_list 데이터를 읽어 YES|NO 판단 후,
                        CSV FILE / XLSX FILE 모두 첫 줄에 더할 수 있어야 한다.
                """
                try:
                    if cell_check == 'YES':
                        def convert_to_alphabet(column_index):

                            # 알파벳 리스트 생성 (A부터 Z까지)
                            alphabet = list(string.ascii_uppercase)

                            # 컬럼 인덱스를 알파벳으로 변환하여 반환
                            if column_index < len(alphabet):
                                return alphabet[column_index]
                            else:
                                # 만약 인덱스가 알파벳 개수를 초과하면, AA, AB, AC와 같이 두 자리 알파벳으로 변환
                                first_letter = alphabet[column_index // len(alphabet) - 1]
                                second_letter = alphabet[column_index % len(alphabet)]
                                return first_letter + second_letter

                        def print_columns_with_alphabet(filename):
                            with open(filename, 'r', encoding='utf-8') as csvfile:
                                reader = csv.reader(csvfile)
                                # 첫 번째 행을 읽어온 후 컬럼의 수를 반환
                                columns = next(reader)
                                result = []
                                for i, column in enumerate(columns):
                                    # 인덱스를 알파벳으로 변환하여 출력
                                    if column == 'TIME':
                                        continue
                                    result.append(convert_to_alphabet(i))
                            return result

                        chunk_size = 1000000  # 100만건이상 케이스 대처
                        dfs = []  # 100만건 안되더라도, 인덱스 요소 추가됨.

                        file_path = csv_file
                        for chunk in pd.read_csv(file_path, chunksize=chunk_size):
                            dfs.append(chunk)

                        # 파일명 처리
                        xlsx_file = csv_file.replace('.csv', '')
                        for df in dfs:
                            print('파일 : ', df)
                            print('길이 : ', len(df))
                        # ExcelWriter 객체 생성 및 xlsxwriter 엔진 사용
                        for i, df in enumerate(dfs):
                            output_file = f'{xlsx_file}_{i + 1}.xlsx'  # 파일 이름 설정
                            xlsx_files.append(output_file)

                            with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
                                # DataFrame을 Excel로 쓰기
                                df.to_excel(writer, sheet_name='Sheet1', index=False)

                                # 워크북과 워크시트 객체 가져오기
                                workbook = writer.book
                                worksheet = writer.sheets['Sheet1']

                                # 조건부 서식 지정
                                color_format = workbook.add_format({'bg_color': '#FFC0CB', 'font_color': '#8B0000'})

                                # 범위 설정 (첫 번째 행은 헤더이므로 2행부터 시작)
                                start_row, _ = 1, 1
                                end_row = len(df)

                                # 범위 알파벳 변환
                                column_range = print_columns_with_alphabet(file_path)

                                # Time 컬럼을 제외한 나머지 컬럼에 조건부 서식 적용
                                for col in column_range:
                                    worksheet.conditional_format(
                                        f'{col}{start_row + 1}:{col}{end_row + 1}',
                                        {'type': 'cell',
                                         'criteria': 'between',
                                         'minimum': 0.5,
                                         'maximum': 1.5,
                                         # 기존에 규칙이 중첩되는 것을 막기 위해 format을 그룹 지음.
                                         'format': color_format})

                                # 변경 사항 저장
                                writer._save()
                                writer.close()
                                print('\t FILE END : ', output_file)
                        print('파일리스트:', xlsx_files)

                except Exception as E:
                    self.message_fail_color = '서식 지정 과정 중 오류가 발생하였습니다.'
                    for i in enumerate(xlsx_files):
                        os.remove(i)

                """ 2024년 5월 22일 CSV 서식 지정 종료 """
                if fixed_check == 'YES':
                    # CSV part
                    with open(csv_file, 'r', newline='') as file:
                        reader = csv.reader(file)
                        rows = list(reader)

                    rows.insert(0, new_data)

                    with open(csv_file, 'w', newline='') as file:
                        writer = csv.writer(file)
                        writer.writerows(rows)

                    # XLSX part
                    for excel_file in xlsx_files:
                        if os.path.exists(excel_file):  # 파일이 존재하는지 확인
                            # 엑셀 파일 열기
                            wb = openpyxl.load_workbook(excel_file)
                            ws = wb.active  # 활성화된 워크시트 선택

                            # 워크시트의 모든 데이터를 읽어옵니다.
                            data = []
                            for row in ws.iter_rows(values_only=True):
                                data.append(row)

                            # 새로운 데이터를 첫 줄에 추가
                            data.insert(0, new_data)

                            # 기존 워크시트를 초기화하고, 수정된 데이터를 다시 씁니다.
                            ws.delete_rows(1, ws.max_row)
                            for row in data:
                                ws.append(row)

                            # 엑셀 파일 저장
                            wb.save(excel_file)
                        else:
                            print(f"파일을 찾을 수 없습니다: {excel_file}")

                print('CSV FILE SUCCESS')

            if mi_dfs:
                mi_result_df = pd.concat(mi_dfs)
                mi_result_df = mi_result_df.drop('Unnamed: 65', axis=1)
                midf1 = mi_result_df
                midf1['Time'] = pd.to_datetime(midf1['Time'], format='%Y%m%d_%H%M%S%f')
                midf1 = midf1.sort_values(by='Time')
                midf1['Time'] = midf1['Time'].dt.strftime('%Y%m%d %H%M%S.%f')
                # mi_result_df = mi_result_df.drop('TemperatureHumidity', axis=1)
                midf1.to_csv(f"{desktop_path}\\{self.get_cpnt_id}_MI_MERGED_{current_date}.csv", index=False)
                path_list.append(f"{desktop_path}\\{self.get_cpnt_id}_MI_MERGED_{current_date}.csv")

            # Excel 실행 / 2024-05-22 중단
            # for path in path_list:
            #     try:
            #         subprocess.Popen(['start', 'excel.exe', path], shell=True)
            #     except Exception as e:
            #         print('열지못함 : ', e)

        except Exception as E:
            self.fail_sig = 1
            print('error : ', E)
            err_msg = traceback.format_exc()
            print(err_msg)
            log_file_path = f'log/{datetime.now().strftime("%Y년%m월%d일%H시%M분%S초")}_CSVerror.log'
            with open(log_file_path, 'w') as log_file:
                log_file.write(str(E))
                log_file.write('self.traceback_str')
                log_file.write(err_msg)
            messagebox.showinfo("오류", f"아래와 같은 오류가 발생하였습니다. log 폴더 내 log 파일을 확인해주십시요.\n{E}")
            messagebox.showinfo("오류", f"가급적 종료하여 주시고, 오류 해결 후 실행해주십시요.")
            self.close_button['state'] = 'normal'

    def open_csv(self):
        thread1 = threading.Thread(target=self.thread_open_csv)
        thread1.start()
        self.execute_translate_button['state'] = 'disabled'

    def thread_open_csv(self):
        try:
            directory_path = self.path
            self.read_csv(directory_path)
            print('read 종료')
            self.merge_csv(directory_path)
            print('merge 종료')

            for filename in os.listdir(directory_path):
                if filename.endswith(".csv") and "_IU_" in filename:
                    file_path = os.path.join(directory_path, filename)
                    os.remove(file_path)
                if filename.endswith(".txt"):
                    file_path = os.path.join(directory_path, filename)
                    os.remove(file_path)
        except Exception as E:
            self.fail_sig = 1
            print('error : ', E)
            err_msg = traceback.format_exc()
            print(err_msg)
            log_file_path = f'log/{datetime.now().strftime("%Y년%m월%d일%H시%M분%S초")}_CSVerror.log'
            with open(log_file_path, 'w') as log_file:
                log_file.write(str(E))
                log_file.write('self.traceback_str')
                log_file.write(err_msg)
            messagebox.showinfo("오류", f"아래와 같은 오류가 발생하였습니다. log 폴더 내 log 파일을 확인해주십시요.\n{E}")
            messagebox.showinfo("오류", "가급적 종료하여 주시고, 오류 해결 후 실행해주십시요.")
            self.close_button['state'] = 'normal'
        print("모든 작업이 완료되었습니다.")
        self.select_folder_button['state'] = 'normal'
        if self.fail_sig == 1:
            self.stop_signal = 1
            messagebox.showerror("실패", f"CSV 변환/결합 실패!\n{self.message_fail_color}")
        elif self.fail_sig == 0:
            self.stop_signal = 1
            messagebox.showinfo("성공",
                                f"CSV 변환/결합 완료!\n소요 시간 : {round(time.time() - self.starting_time)}초\n{self.message_fail_color}")
        self.file_tree2.delete("2")
        self.hide_loading()

    def refresh_connection(self):
        global mysql_connect
        global influx_connect
        # MySQL 연결 설정
        mysql_config = self.read_db_config(self.connect_ini_file_name, 'MySQL')
        mysql_connect = mysql.connector.connect(
            host=mysql_config['host'],
            port=mysql_config['port'],
            user=mysql_config['user'],
            password=mysql_config['password'],
            database=mysql_config['database'],
            allow_local_infile=True,
            autocommit=False
        )

        # InfluxDB 연결 설정
        influx_config = self.read_db_config(self.connect_ini_file_name, 'InfluxDB')
        influx_host = influx_config['host']
        influx_port = int(influx_config['port'])
        influx_dbname = influx_config['dbname']
        influx_connect = InfluxDBClient(influx_host, influx_port, database=influx_dbname)
        print('새로고침 완료')

    # InfluxDB 를 select 한 결과를 현재 경로\output.csv 에 생성
    def make_csv(self, influx_measurement):

        self.refresh_connection()
        query = f'SELECT * FROM {influx_measurement} ORDER BY time DESC LIMIT 1'
        result = influx_connect.query(query)
        columns = result.raw['series'][0]['columns']
        latest_data = list(result.raw['series'][0]['values'])[0]
        latest_time_str = latest_data[0]
        latest_time = datetime.strptime(latest_time_str, '%Y-%m-%dT%H:%M:%S.%fZ')
        seven_days_ago = latest_time - timedelta(days=7)

        # 위 과정은 날짜 타입으로 formatting 한 것이며, 이를 통해 나온 변수 'sevne_days_ago' 를 통해
        # query 를 재생성
        query = f'SELECT * FROM {influx_measurement} WHERE time > \'{seven_days_ago.isoformat()}Z\''
        result = influx_connect.query(query)
        with open(f'{influx_measurement}_output.csv', mode='w', newline='') as file:
            writer = csv.writer(file)
            writer.writerow(columns)
            for row in result.raw['series'][0]['values']:
                formatted_date = datetime.strptime(row[0].replace('Z', '').split('.')[0], '%Y-%m-%dT%H:%M:%S').strftime(
                    '%Y-%m-%d %H:%M:%S')
                formatted_row = [formatted_date] + row[1:]
                writer.writerow(formatted_row)
        csv_file = influx_measurement + "_output.csv"
        df = pd.read_csv(csv_file)
        df.fillna('NULL', inplace=True)
        df.to_csv(csv_file, index=False)
        print('csv 파일 생성 완료')
        self.insert_csv(csv_file, influx_measurement, columns)
        influx_connect.close()

    # LOAD DATA INFILE 을 LOCAL 타입으로 설정한다. ( client side 에서 파일 insert 가능 )
    def insert_csv(self, file_name, tb_name, columns):
        if tb_name == 'iu_data':
            tb_name = 'tb_eqp_iu'
        elif tb_name == 'mi_data':
            tb_name = 'tb_eqp_mi'
        print(file_name, ' 파일명 ', tb_name, ' 테이블명 ', columns, ' 컬럼명 ')

        # 칼럼 목록을 문자열로 만듭니다.
        set_columns = ', '.join(f"@`{col}`" if col != "eqp_id" else f"{col}" for col in columns)
        print(set_columns)
        # SET 부분을 동적으로 생성합니다.
        set_part = ''
        for col in columns:
            if col != 'eqp_id':
                if col == 'Time':  # 날짜타입으로 안맞춰도 들어가긴 함
                    print()
                    # set_part += f"{col} = STR_TO_DATE(@{col}, '%Y-%m-%d %H:%i:%s'), "
                else:
                    # 특수 문자가 포함된 컬럼 이름을 처리합니다.
                    col_name = f"{col}"
                    set_part += f"`{col_name}` = @`{col}`, "
        print(set_part)
        set_part = set_part.rstrip(', ')  # 콤마 빼면 들어가는 이슈있음
        cursor = mysql_connect.cursor()
        load_sql = f"""
                LOAD DATA LOCAL INFILE '{file_name}'
                INTO TABLE {tb_name}
                FIELDS TERMINATED BY ','
                ENCLOSED BY '"'
                LINES TERMINATED BY '\n'
                IGNORE 1 ROWS
                ({set_columns})
                SET {set_part};
        """
        cursor.execute(load_sql)
        mysql_connect.commit()
        cursor.close()
        mysql_connect.close()
        print('csv 파일 삽입 완료')

    def execute_insert(self):
        result = messagebox.askquestion("경고", "예상 시간이 6분이상 소요되는 작업입니다.\n계속하시겠습니까?")
        if result == "yes":
            # Yes를 선택한 경우 추가 작업 수행
            thread2 = threading.Thread(target=self.animations)
            thread2.start()
            thread1 = threading.Thread(target=self.execute_thread)
            thread1.start()
            self.execute_insert_button['state'] = 'disabled'
            self.file_tree2.insert('', 'end', iid="3", values=('최근 7일치 데이터 삽입',))
        else:
            # No를 선택한 경우 추가 작업 수행
            pass

    def execute_thread(self):
        self.refresh_connection()
        influx_measurement = 'mi_data'  # influx 테이블명
        self.make_csv(influx_measurement)

        self.refresh_connection()
        influx_measurement = 'iu_data'  # influx 테이블명
        self.make_csv(influx_measurement)

        print("모든 작업이 완료되었습니다.")
        self.execute_insert_button['state'] = 'normal'
        messagebox.showinfo("Complete", f"데이터 업로드 완료!\n총 소요 시간 : {time.time() - self.stime}초")
        self.file_tree2.delete("3")

    def animations(self):

        self.stime = time.time()
        self.progressing()
        self.animate_loading1()

    def main(self, input_directory):
        try:
            def convert_to_datetime(timestamp):
                timestamp_str = str(timestamp)
                timestamp_str = timestamp_str[:13]

                timestamp_int = timestamp_str[7:]
                timestamp_abc = timestamp_int[3:6] + '000'

                seconds = int(timestamp_str[:10])
                microseconds = int(timestamp_abc)

                dt = datetime.fromtimestamp(seconds)
                dt += timedelta(microseconds=microseconds)

                formatted_dt = dt.strftime('%Y-%m-%d %H:%M:%S.%f')

                return formatted_dt

            def time_translate(input_file, output_file):
                with open(input_file, 'r') as f:
                    lines = f.readlines()

                with open(output_file, 'w') as f:
                    for line in lines:
                        unix_timestamp_match = re.search(r'\b\d{19}\b', line)
                        if unix_timestamp_match:
                            unix_timestamp = int(unix_timestamp_match.group())
                            korean_time = convert_to_datetime(unix_timestamp)
                            line = line.replace(str(unix_timestamp), korean_time)
                            f.write(line)
                        else:
                            print('탐지불가')


            value = input_directory

            delete_files = []
            for file in os.listdir(input_directory):
                if 'IU' in file and file.endswith('.txt'):
                    file_path = os.path.join(input_directory, file)
                    if os.path.isfile(file_path):
                        with open(file_path, 'r', encoding='latin-1') as f:
                            content = f.read()
                            if re.search(r'\b\d{19}\b', content):
                                logging.info(file_path)
                            else:
                                delete_files.append(file_path)

            for files in delete_files:
                os.remove(files)

            def merge_file(input_file, output_queue):
                with open(input_file, 'r') as infile:
                    content = infile.read()
                    output_queue.put(content)

            output_file = os.path.join(value, 'output_merge.txt')
            output_queue = queue.Queue()

            input_files = [os.path.join(input_directory, file) for file in os.listdir(input_directory) if
                           file.endswith('.txt')]

            with concurrent.futures.ThreadPoolExecutor() as executor:
                futures = [executor.submit(merge_file, file, output_queue) for file in input_files]
                concurrent.futures.wait(futures)

            with open(output_file, 'w') as outfile:
                while not output_queue.empty():
                    outfile.write(output_queue.get())

            file_path = os.path.join(value, 'output_merge.txt')
            temp_file_path = file_path + ".tmp"

            with open(file_path, 'r', encoding='utf-8') as f:
                with open(temp_file_path, 'w', encoding='utf-8') as temp_f:
                    for line in f:
                        if re.search(r'\b\d{19}\b', line):
                            temp_f.write(line)

            os.remove(file_path)
            os.rename(temp_file_path, file_path)

            def delete_lines(input_directory, output_directory):
                with open(input_directory, "r") as f:
                    lines = f.readlines()

                # 출력할 행을 담을 리스트
                output_lines = []

                # 각 행을 반복하여 확인
                for line in lines:
                    if "SR_" in line or "TR_" in line or "HDLC_K" in line:
                        output_lines.append(line)

                # 출력 파일에 쓰기
                with open(output_directory, "w") as f:
                    f.writelines(output_lines)

            delete_lines(os.path.join(value, 'output_merge.txt'), os.path.join(value, 'output_merge.txt'))

            # 시간이 걸리더라도 해야함. 개선 가능성은 ?
            time_translate(file_path, output_file=os.path.join(value, 'output_array.txt'))

            def rearrange_strings(string_list):
                timestamped_strings = [(re.findall(r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}.\d{6}', string)[0], string) for
                                       string in string_list]
                sorted_strings = sorted(timestamped_strings, key=lambda x: x[0])
                return [string for _, string in sorted_strings]

            with open(os.path.join(value, 'output_array.txt'), 'r') as file:
                string_list = file.read().splitlines()

            result = rearrange_strings(string_list)

            with open(os.path.join(value, 'output_array.txt'), 'w') as file:
                for string in result:
                    file.write(string + '\n')

            # 시간 변환하고, 날짜별로 끊어야 아래 로직이 날짜별 운용 이력을 잡을 수 있음 개선 가능성은 ?
            def process_line_date(line):
                # 정규표현식 패턴 정의
                date_pattern = re.compile(r'(\d{4}-\d{2}-\d{2})')

                # 각 줄에서 날짜를 추출
                match = date_pattern.search(line)
                if match:
                    date = match.group(1)
                    return date, line
                return None, None

            def date_collect(value, file_path):
                # 'outputs' 폴더 생성
                output_directory = os.path.join(value, 'data by date')
                os.makedirs(output_directory, exist_ok=True)

                # 날짜별로 데이터를 저장할 defaultdict 생성
                data_by_date = defaultdict(list)

                # 텍스트 파일 읽기
                with open(file_path, 'r') as file:
                    # concurrent.futures를 사용하여 멀티스레딩 구현
                    with concurrent.futures.ThreadPoolExecutor() as executor:
                        results = list(executor.map(process_line_date, file))

                # 결과를 처리하여 데이터를 날짜별로 저장
                for date, line in results:
                    if date:
                        data_by_date[date].append(line)

                # 각 날짜별로 파일 생성하여 데이터 쓰기
                for date, data in data_by_date.items():
                    file_name = os.path.join(output_directory, f'{date.replace("-", "_")}.txt')
                    with open(file_name, 'w') as output_file:
                        output_file.writelines(data)

            input_directory = os.path.join(value, 'output_array.txt')
            date_collect(value, input_directory)

            def find_and_print_find(pattern, lines, output_file):
                found_first = False
                for line in lines:
                    if re.search(pattern, line):
                        output_file.write(line.rstrip() + '\n')
                        found_first = True
                        break
                if not found_first:
                    output_file.write('\n')

                found_last = False
                for line in reversed(lines):
                    if re.search(pattern, line):
                        output_file.write(line.rstrip() + '\n')
                        found_last = True
                        break
                if not found_last:
                    output_file.write('\n')

            def process_files(input_directory, output_file_path):
                with open(output_file_path, 'a') as output_file:
                    for filename in os.listdir(input_directory):
                        if filename.endswith(".txt"):
                            input_file_path = os.path.join(input_directory, filename)
                            with open(input_file_path, 'r') as input_file:
                                lines = input_file.readlines()
                                find_and_print_find(re.compile(r'SR_'), lines, output_file)
                                find_and_print_find(re.compile(r'TR_'), lines, output_file)
                                find_and_print_find(re.compile(r'HDLC_K'), lines, output_file)
                                output_file.write('\n')

            # keywords = ['SR_', 'TR_', 'HDLC_']
            input_directory = os.path.join(value, 'data by date')
            output_file_path = os.path.join(value, 'output_works.txt')
            process_files(input_directory, output_file_path)

            # 입력 파일과 출력 파일 지정
            input_file = os.path.join(value, 'output_works.txt')
            output_file = os.path.join(value, 'output_result.txt')

            with open(input_file, 'r') as f:
                lines = f.readlines()

            with open(output_file, 'w') as f:
                for line in lines:
                    unix_timestamp_match = re.search(r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}.\d{6}', line)
                    if unix_timestamp_match:
                        line = line.replace('iu_data,eqp_id=', '')
                        if 'SR_' in line:
                            line = re.sub(r'SR_.*? ', 'SR ', line)
                        if 'TR_' in line:
                            line = re.sub(r'TR_.*? ', 'TR ', line)
                        if 'HDLC_' in line:
                            line = re.sub(r'HDLC_.*? ', 'RCS ', line)
                        f.write(line)

            input_file = os.path.join(value, 'output_result.txt')
            output_file = os.path.join(value, 'output_working.txt')
            sr_list = []
            tr_list = []
            rcs_list = []
            pattern = r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}'
            with open(output_file, 'w') as result_file:
                with open(input_file, 'r') as file:
                    for line in file:
                        if re.search(r'SR', line):
                            sr_list.append(line.strip())
                            if len(sr_list) == 2:
                                match0 = re.search(r'KSAM\d+ [A-Z]+ \d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}', sr_list[0])
                                match1 = re.search(pattern, sr_list[-1])
                                result = match0.group() + ' ' + match1.group() + '\n'
                                result_file.write(result)
                                sr_list = []
                        if re.search(r'TR', line):
                            tr_list.append(line.strip())
                            if len(tr_list) == 2:
                                match0 = re.search(r'KSAM\d+ [A-Z]+ \d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}', tr_list[0])
                                match1 = re.search(pattern, tr_list[-1])
                                result = match0.group() + ' ' + match1.group() + '\n'
                                result_file.write(result)
                                tr_list = []
                        if re.search(r'RCS', line):
                            rcs_list.append(line.strip())
                            if len(rcs_list) == 2:
                                match0 = re.search(r'KSAM\d+ [A-Z]+ \d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}', rcs_list[0])
                                match1 = re.search(pattern, rcs_list[-1])
                                result = match0.group() + ' ' + match1.group() + '\n'
                                result_file.write(result)
                                rcs_list = []

            output_file_path = os.path.join(value, 'output_array.txt')

            def rearrange_strings(string_list):
                timestamped_strings = [(re.findall(r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}.\d{6}', string)[0], string) for
                                       string in string_list]
                sorted_strings = sorted(timestamped_strings, key=lambda x: x[0])
                return [string for _, string in sorted_strings]

            with open(output_file_path, 'r') as file:
                string_list = file.read().splitlines()

            result = rearrange_strings(string_list)

            with open(output_file_path, 'w') as file:
                for string in result:
                    file.write(string + '\n')

            def process_single_file_SR(input_file_path, output_file_path):
                with open(output_file_path, 'w') as output_file:
                    keywords = [
                        r'SR_RACED1_RADAFAL',
                        r'SR_RACED1_VSWRFAL',
                        r'SR_RACED1_TWTAFAL',
                        r'SR_RACED1_INTEFAL',
                        r'SR_RACED1_ANTEFAL',
                        r'SR_RACED1_RPUFAL',
                        r'SR_RACED1_TXRXFAL'
                    ]
                    with open(input_file_path, 'r') as file:
                        lines = file.readlines()
                        for line in lines:
                            for keyword in keywords:
                                if re.search(keyword, line):
                                    line = line.replace('iu_data,eqp_id=', '')
                                    output_file.write(line.rstrip() + '\n')
                                    break

                input_directory = output_file_path
                output_directory = os.path.join(value, 'output_sr_result.txt')
                str_list = []
                match_list = []
                with open(input_directory, 'r') as file:
                    lines = file.readlines()
                lines = [line.strip() for line in lines if line.strip()]
                with open(input_directory, 'w') as file:
                    file.write('\n'.join(lines))

                # Initialize variables
                TXRXFAL = True
                RPUFA = True
                ANTEFAL = True
                INTEFAL = True
                TWTAFAL = True
                VSWRFAL = True
                RADAFAL = True
                anyonce = True
                # 0.*?
                patterns = [
                    r'SR_RACED1_TXRXFAL=0.*?',
                    r'SR_RACED1_RPUFAL=0.*?',
                    r'SR_RACED1_ANTEFAL=0.*?',
                    r'SR_RACED1_INTEFAL=0.*?',
                    r'SR_RACED1_TWTAFAL=0.*?',
                    r'SR_RACED1_VSWRFAL=0.*?',
                    r'SR_RACED1_RADAFAL=0.*?',
                    # 1~7
                    r'SR_RACED1_TXRXFAL=1.*?',
                    r'SR_RACED1_RPUFAL=1.*?',
                    r'SR_RACED1_ANTEFAL=1.*?',
                    r'SR_RACED1_INTEFAL=1.*?',
                    r'SR_RACED1_TWTAFAL=1.*?',
                    r'SR_RACED1_VSWRFAL=1.*?',
                    r'SR_RACED1_RADAFAL=1.*?'
                    # 8~14
                ]
                with open(input_directory, "r") as file:
                    lines = file.readlines()

                def resize(str1, str2, str3, end):
                    match0 = re.search(r'KSAM\d+ ', str1)
                    match1 = re.search(r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}.\d{6}', str1)
                    match2 = re.search(r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}.\d{6}', str2)
                    match_append()
                    match3 = ', '.join(map(str, str3))
                    print('str1', str1)
                    print('match1', match1)
                    print('str2', str2)
                    print('match2', match2)
                    print('str3', str3)
                    print('match3', match3)
                    print(match0.group() + match3 + ' ' + match2.group() + ' ' + match1.group())
                    print(' 지점', end)
                    if end == 0:
                        output_file.write(
                            match0.group() + match3 + ' ' + match2.group() + ' ' + match1.group() + ' N' + '\n')
                    elif end == 1:
                        output_file.write(
                            match0.group() + match3 + ' ' + match2.group() + ' ' + match1.group() + ' Y' + '\n')

                def match_append():
                    for data in str_list:
                        match = re.search(patterns[7], data)
                        if match and match.group() not in match_list:
                            match_list.append(match.group())
                        match = re.search(patterns[8], data)
                        if match and match.group() not in match_list:
                            match_list.append(match.group())
                        match = re.search(patterns[9], data)
                        if match and match.group() not in match_list:
                            match_list.append(match.group())
                        match = re.search(patterns[10], data)
                        if match and match.group() not in match_list:
                            match_list.append(match.group())
                        match = re.search(patterns[11], data)
                        if match and match.group() not in match_list:
                            match_list.append(match.group())
                        match = re.search(patterns[12], data)
                        if match and match.group() not in match_list:
                            match_list.append(match.group())
                        match = re.search(patterns[13], data)
                        if match and match.group() not in match_list:
                            match_list.append(match.group())

                with open(output_directory, "w") as output_file:  # 새로운 파일을 쓰기 모드로 열기
                    for i, line in enumerate(lines):
                        if re.search(patterns[0], line):
                            TXRXFAL = True
                        if re.search(patterns[1], line):
                            RPUFA = True
                        if re.search(patterns[2], line):
                            ANTEFAL = True
                        if re.search(patterns[3], line):
                            INTEFAL = True
                        if re.search(patterns[4], line):
                            TWTAFAL = True
                        if re.search(patterns[5], line):
                            VSWRFAL = True
                        if re.search(patterns[6], line):
                            RADAFAL = True
                        if re.search(patterns[7], line):
                            TXRXFAL = False
                            anyonce = False
                            str_list.append(line.strip())
                        if re.search(patterns[8], line):
                            RPUFA = False
                            anyonce = False
                            str_list.append(line.strip())
                        if re.search(patterns[9], line):
                            ANTEFAL = False
                            anyonce = False
                            str_list.append(line.strip())
                        if re.search(patterns[10], line):
                            INTEFAL = False
                            anyonce = False
                            str_list.append(line.strip())
                        if re.search(patterns[11], line):
                            TWTAFAL = False
                            anyonce = False
                            str_list.append(line.strip())
                        if re.search(patterns[12], line):
                            VSWRFAL = False
                            anyonce = False
                            str_list.append(line.strip())
                        if re.search(patterns[13], line):
                            RADAFAL = False
                            anyonce = False
                            str_list.append(line.strip())

                        if not anyonce:
                            #
                            if TXRXFAL and RPUFA and ANTEFAL and INTEFAL and TWTAFAL and VSWRFAL and RADAFAL:
                                str_list.append(line.strip())
                                # print('linestrip', line.strip())
                                resize(str_list[-1], str_list[0], match_list, 0)
                                match_list = []
                                str_list = []
                                anyonce = True
                            # 마지막 행 처리
                            if i == len(lines) - 1:
                                if not (TXRXFAL and RPUFA and ANTEFAL and INTEFAL and TWTAFAL and VSWRFAL and RADAFAL):
                                    str_list.append(lines[len(lines) - 1].strip())
                                    # print('lines[len(lines) - 2].strip()', lines[len(lines) - 1].strip())
                                    resize(str_list[-1], str_list[0], match_list, 1)
                                    match_list = []
                                    str_list = []

                input_file_path = os.path.join(value, 'output_sr_result.txt')
                output_file_path = os.path.join(value, 'output_sr_result.txt')
                processed_lines = []

                # 파일을 읽어들임
                with open(input_file_path, 'r') as file:
                    # 파일의 각 라인에 대해 반복
                    for line in file:
                        # 'HDLC_K,' 또는 'HDLC_K ' 또는 'TR_,' 또는 'TR_ ' 또는 'SR_,' 또는 'SR_ '을 삭제하고 해당 자리에 공백을 추가
                        processed_line = line.replace('=1', '').replace(',', '')
                        # processed_line = line.replace('=1.000000', '').replace(',', '')

                        # KSAM 다음에 오는 숫자를 찾아 대체
                        processed_line = re.sub(r'(KSAM)(\d+)', r'\1\2 SR', processed_line)

                        processed_lines.append(processed_line)

                # 처리된 결과를 파일에 쓰기
                with open(output_file_path, 'w') as output_file:
                    for line in processed_lines:
                        output_file.write(line)

            def process_single_file_TR(input_file_path, output_file_path):
                with open(output_file_path, 'w') as output_file:
                    keywords = [
                        r'TR_KUMEL1_AVARAD',
                        r'TR_KUMSD1_SYNCRO',
                        r'TR_KUMEL1_TEMPER'
                    ]
                    with open(input_file_path, 'r') as file:
                        lines = file.readlines()
                        for line in lines:
                            for keyword in keywords:
                                if re.search(keyword, line):
                                    line = line.replace('iu_data,eqp_id=', '')
                                    output_file.write(line.rstrip() + '\n')
                                    break

                input_directory = output_file_path
                output_directory = os.path.join(value, 'output_tr_result.txt')

                str_list = []
                match_list = []
                with open(input_directory, 'r') as file:
                    lines = file.readlines()
                lines = [line.strip() for line in lines if line.strip()]
                with open(input_directory, 'w') as file:
                    file.write('\n'.join(lines))

                # Initialize variables
                Syncro = True
                Avarad = True
                Temperr = True
                anyonce = True

                patterns = [
                    r'TR_KUMSD1_SYNCRO=0.*?',
                    r'TR_KUMEL1_AVARAD=0.*?',
                    r'TR_KUMEL1_TEMPER=0.*?',
                    r'TR_KUMSD1_SYNCRO=1.*?',
                    r'TR_KUMEL1_AVARAD=1.*?',
                    r'TR_KUMEL1_TEMPER=1.*?'
                ]
                with open(input_directory, "r") as file:
                    lines = file.readlines()

                def match_append():
                    for data in str_list:
                        match = re.search(patterns[3], data)
                        if match and match.group() not in match_list:
                            match_list.append(match.group())
                        match = re.search(patterns[4], data)
                        if match and match.group() not in match_list:
                            match_list.append(match.group())
                        match = re.search(patterns[5], data)
                        if match and match.group() not in match_list:
                            match_list.append(match.group())

                def resize(str1, str2, str3, end):
                    match0 = re.search(r'KSAM\d+ ', str1)
                    match1 = re.search(r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}.\d{6}', str1)
                    match2 = re.search(r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}.\d{6}', str2)
                    match_append()
                    match3 = ', '.join(map(str, str3))
                    # print(' 지점', end)
                    # print('결과', match0.group() + match3 + ' ' + match2.group() + ' ' + match1.group())
                    if end == 0:
                        output_file.write(
                            match0.group() + match3 + ' ' + match2.group() + ' ' + match1.group() + ' N' + '\n')
                    elif end == 1:
                        output_file.write(
                            match0.group() + match3 + ' ' + match2.group() + ' ' + match1.group() + ' Y' + '\n')

                with open(output_directory, "w") as output_file:  # 새로운 파일을 쓰기 모드로 열기
                    for i, line in enumerate(lines):
                        if re.search(patterns[0], line):
                            Syncro = True
                        if re.search(patterns[1], line):
                            Avarad = True
                        if re.search(patterns[2], line):
                            Temperr = True
                        if re.search(patterns[3], line):
                            Syncro = False
                            anyonce = False
                            str_list.append(line.strip())
                        if re.search(patterns[4], line):
                            Avarad = False
                            anyonce = False
                            str_list.append(line.strip())
                        if re.search(patterns[5], line):
                            Temperr = False
                            anyonce = False
                            str_list.append(line.strip())

                        if not anyonce:
                            if Syncro and Avarad and Temperr:
                                str_list.append(line.strip())
                                resize(str_list[-1], str_list[0], match_list, 0)
                                match_list = []
                                str_list = []
                                anyonce = True
                            if i == len(lines) - 1:
                                if Syncro == False or Avarad == False or Temperr == False:
                                    str_list.append(lines[len(lines) - 1].strip())
                                    resize(str_list[-1], str_list[0], match_list, 1)
                                    match_list = []
                                    str_list = []

                input_file_path = os.path.join(value, 'output_tr_result.txt')
                output_file_path = os.path.join(value, 'output_tr_result.txt')

                processed_lines = []

                # 파일을 읽어들임
                with open(input_file_path, 'r') as file:
                    # 파일의 각 라인에 대해 반복
                    for line in file:
                        # 'HDLC_K,' 또는 'HDLC_K ' 또는 'TR_,' 또는 'TR_ ' 또는 'SR_,' 또는 'SR_ '을 삭제하고 해당 자리에 공백을 추가
                        processed_line = line.replace('=1', '').replace(',', '')
                        # processed_line = line.replace('=1.000000', '').replace(',', '')

                        processed_line = re.sub(r'(KSAM)(\d+)', r'\1\2 TR', processed_line)
                        # print('결과 : ', processed_line)
                        processed_lines.append(processed_line)

                # 처리된 결과를 파일에 쓰기
                with open(output_file_path, 'w') as output_file:
                    # print(processed_lines)
                    for line in processed_lines:
                        output_file.write(line)

            ind = os.path.join(value, 'output_array.txt')
            outd = os.path.join(value, 'output_sr_error.txt')
            process_single_file_SR(ind, outd)

            ind = os.path.join(value, 'output_array.txt')
            outd = os.path.join(value, 'output_tr_error.txt')
            process_single_file_TR(ind, outd)
        except Exception as E:
            # TraceBack : traceback.print_exc()
            print('에러 유형 : ', str(E))

    def working_query(self, value, error_path, result_path):
        try:
            mysql_config = self.read_db_config(self.connect_ini_file_name, 'MySQL')
            connection = mysql.connector.connect(
                host=mysql_config['host'],
                port=mysql_config['port'],
                user=mysql_config['user'],
                password=mysql_config['password'],
                database=mysql_config['database']

            )

            # 운용 이력 insert
            print('working query 시작합니다 =================')

            # MySQL 커서 생성
            # try:
            cursor = connection.cursor()
            print('결과11 : ', cursor)
            # 파일 열기
            input_directory = os.path.join(value, 'output_working.txt')
            with open(input_directory, 'r') as file:
                # 모든 데이터를 담을 리스트 초기화
                working_history = []

                # 파일에서 모든 줄을 읽어와서 리스트에 추가
                for line in file:
                    data = line.strip().split(' ')
                    eqp_id = data[0]
                    cpnt_id = data[1]
                    work_date = data[2].replace('-', '')
                    work_stime = data[3].replace(':', '')
                    work_etime = data[5].replace(':', '')
                    wk_sdt = data[2] + ' ' + data[3]
                    wk_edt = data[4] + ' ' + data[5]
                    # 각 행의 데이터를 튜플로 만들어 리스트에 추가
                    working_history.append((eqp_id, cpnt_id, work_date, work_stime, work_etime, wk_sdt, wk_edt))

                # 프로시저 호출 및 파라미터 설정
                cursor.executemany("CALL cbmplus.sp_equip_working_history(%s, %s, %s, %s, %s, %s, %s)",
                                   working_history)
                
            print(error_path, result_path)

            with open(error_path, 'r') as file:

                first_line = file.readline().strip()

                data = first_line.split(' ')
                print('오류 확인', data)
                dataa = data
                eqp_id = data[0]
                if 'sr' in error_path:
                    cpnt_id = 'SR'
                    # bd_till = None
                elif 'tr' in error_path:
                    cpnt_id = 'TR'

                bd_sdt = dataa[-2] + ' ' + dataa[-1]

                print(eqp_id, cpnt_id, bd_sdt)
                sql_query = f"SELECT bd_sdt, bd_edt, bd_still FROM tb_equip_breakdown WHERE eqp_id = '{eqp_id}' and cpnt_id = '{cpnt_id}' and bd_sdt < '{bd_sdt}' and bd_edt is null order by bd_sdt desc;"
                cursor.execute(sql_query)
                result = cursor.fetchall()
                connection.commit()
                print('쿼리결과', result)

                # 해당 데이터가 있는 경우 : NULL 이 아닌 경우
                # 데이터 모두 insert
                if len(result) == 0:
                    print('고장이 켜진 데이터가 없음')
                    print('결과 : ', self.insert_query(result_path))
                    cursor.executemany("CALL cbmplus.sp_equip_error_history(%s, %s, %s, %s, %s, %s, %s, %s)",
                                       self.insert_query(result_path))
                    connection.commit()

                # 해당 데이터가 없는 경우 : NULL 인 경우
                # 데이터 판독해야함
                elif len(result) > 0:
                    print('고장이 켜진 데이터가 존재함')
                    if 'tr' in error_path:
                        print('SR 데이터입니다.')
                        # 모든 줄이 1인 경우를 검사, 데이터를 집어넣지않는다.

                        if self.check_tr(error_path) is True:
                            print('모든 데이터 고장, 데이터 INSERT X')
                        # 고장이 끝난 지점이 존재하는 경우
                        elif self.check_tr(error_path) is not True:
                            print('고장이 끝난 지점이 존재합니다. update 하고 나머지 데이터 insert')
                            print('고장이 끝난 지점', self.check_tr(error_path))
                            ftime = re.search(r"\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}.\d{6}",
                                              self.check_tr(error_path)).group()
                            print(ftime)
                            sql_query = f"UPDATE tb_equip_breakdown SET modr_id = 'SYSTEM', mod_dt = NOW(), bd_edt = '{ftime}', rp_edt = '{ftime}' WHERE eqp_id = '{eqp_id}' AND cpnt_id = '{cpnt_id}' AND bd_sdt = '{result[0][0]}';"
                            cursor.execute(sql_query)
                            connection.commit()
                            for idx, row in enumerate(self.insert_query(result_path)):
                                if ftime in row:
                                    print('해당 row는 제외합니다.')
                                else:
                                    print('다른 데이터')
                                    data = [tuple(row)]
                                    cursor.executemany(
                                        "CALL cbmplus.sp_equip_error_history(%s, %s, %s, %s, %s, %s, %s, %s)", data)
                                    connection.commit()

                    elif 'sr' in error_path:
                        print('SR 데이터입니다.')
                        with open(error_path, 'r') as file:
                            first_line = file.readline().strip()

                            # 첫 줄 시간 데이터
                            ftime = re.search(r"\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}.\d{6}", first_line).group()

                            # 모든 줄이 1인 경우를 검사, 데이터를 집어넣지않는다.
                            if self.check_lines(error_path):
                                print('모든 줄이 =1 이 포함됩니다.')

                            else:
                                # 첫 줄이 1인 경우
                                print('모든 줄이 =1 이 아닙니다..')
                                if '=1' in first_line:
                                    # 해당 ftime 을 쿼리에 던지지않고, sr_result.txt 을 통해 ftime 과 겹치는 행을 찾아 0 이 되는 시점을 update 그리고 나머지 행들 모두 insert

                                    for idx, row in enumerate(self.insert_query(result_path)):
                                        if ftime in row:
                                            data = list(row)
                                            data = data[-2]

                                            sql_query = f"UPDATE tb_equip_breakdown SET modr_id = 'SYSTEM', mod_dt = NOW(), bd_edt = '{data}', rp_edt = '{data}' WHERE eqp_id = '{eqp_id}' AND cpnt_id = '{cpnt_id}' AND bd_sdt = '{result[0][0]}';"
                                            cursor.execute(sql_query)
                                            connection.commit()

                                        else:
                                            data = [tuple(row)]
                                            cursor.executemany(
                                                "CALL cbmplus.sp_equip_error_history(%s, %s, %s, %s, %s, %s, %s, %s)",
                                                data)
                                            connection.commit()

                                # 첫 줄이 0인 경우
                                else:
                                    print('아니')
                                    # 해당 ftime 을 쿼리에 던진다. 그리고 update. 그리고 나머지 sr_result.txt 행들 모두 insert
                                    sql_query = f"UPDATE tb_equip_breakdown SET modr_id = 'SYSTEM', mod_dt = NOW(), bd_edt = '{ftime}', rp_edt = '{ftime}' WHERE eqp_id = '{eqp_id}' AND cpnt_id = '{cpnt_id}' AND bd_sdt = '{result[0][0]}';"
                                    cursor.execute(sql_query)
                                    connection.commit()
                                    cursor.executemany(
                                        "CALL cbmplus.sp_equip_error_history(%s, %s, %s, %s, %s, %s, %s, %s)",
                                        self.insert_query(result_path))
                                    self.connection.commit()
        except Exception as err:
            if 'Duplicate' and 'index' not in str(err):
                pass
        finally:
            # 연결 및 커서 닫기
            cursor.close()
            connection.close()

    def check_lines(self, error_path):
        with open(error_path, 'r') as file:
            for line in file:
                if '=1' not in line:
                    return False
        return True

    def insert_query(self, result_path):
        bd_till = ''
        if 'tr' in result_path:
            print()
            # bd_till = ''.join(still_bad(result_path))
            # bd_till = list(bd_till)

        with open(result_path, 'r') as file:
            bad_history = []
            for line in file:
                data = line.strip().split(' ')
                eqp_id = data[0]
                cpnt_id = data[1]
                bd_sdt = data[-5] + ' ' + data[-4]
                rp_sdt = data[-5] + ' ' + data[-4]
                rp_type = ','.join(data[2:-5])
                check_type = data[-1]
                if check_type == 'Y':
                    bd_edt = None
                    rp_edt = None
                    bad_history.append((eqp_id, cpnt_id, bd_sdt, bd_edt, bd_till, rp_sdt, rp_edt, rp_type))
                elif check_type == 'N':
                    rp_edt = data[-3] + ' ' + data[-2]
                    bd_edt = data[-3] + ' ' + data[-2]
                    bad_history.append((eqp_id, cpnt_id, bd_sdt, bd_edt, None, rp_sdt, rp_edt, rp_type))

            return bad_history

    def still_bad(self, file_path):
        last_line = ''
        bd_till = []
        with open(file_path, 'r') as file:
            for line in file:
                if line.strip():
                    last_line = line
        if 'Y' in last_line:
            if 'SYNCRO' in last_line:
                bd_till.append('')
            if 'TEMPER' in last_line:
                bd_till.append('')
            if 'AVARAD' in last_line:
                bd_till.append('')
        else:
            print()
        return bd_till

    def check_tr(self, filename):
        # 전역 변수 설정
        syncro_found = False
        temper_found = False
        avarad_found = False

        # 주어진 표현식들
        syncro_regex = re.compile(r'SYNCRO=0.*?')
        temper_regex = re.compile(r'TEMPER=0.*?')
        avarad_regex = re.compile(r'AVARAD=0.*?')

        # 파일 열기
        with open(filename, 'r') as file:
            error_t = False
            # 각 라인에 대해 반복
            for line in file:
                # SYNCRO 표현식 매치 여부 확인
                if syncro_regex.search(line):
                    syncro_found = True
                # TEMPER 표현식 매치 여부 확인
                if temper_regex.search(line):
                    temper_found = True
                # AVARAD 표현식 매치 여부 확인
                if avarad_regex.search(line):
                    avarad_found = True

                if 'SYNCRO' in line:
                    if not syncro_regex.search(line):
                        syncro_found = False

                if 'TEMPER' in line:
                    # TEMPER 표현식 매치 여부 확인
                    if not temper_regex.search(line):
                        temper_found = False

                if 'AVARAD' in line:
                    # AVARAD 표현식 매치 여부 확인
                    if not avarad_regex.search(line):
                        avarad_found = False

                # 모든 조건을 만족하는 경우 해당 라인 출력 후 종료
                if syncro_found and temper_found and avarad_found:
                    normal_line = line.strip()
                    return normal_line
                else:
                    error_t = True
            return error_t

    def create_timestamped_directory(self, directory):
        # 현재 시간을 기준으로 폴더 이름 생성 (예: 2022-03-30_12-30-45)
        current_time = datetime.now().strftime("%Y년%m월%d일%H시%M분%S초_결과")
        timestamped_directory = os.path.join(directory, current_time)
        os.makedirs(timestamped_directory, exist_ok=True)
        return timestamped_directory

    def move_files_to_directory(self, files_to_move, destination_directory):
        for file_path in files_to_move:
            file_name = os.path.basename(file_path)
            shutil.move(file_path, os.path.join(destination_directory, file_name))
            print(f"Moved '{file_path}' to '{destination_directory}'")

    def delete_directory(self, list_to_delete):
        for path in list_to_delete:
            if os.path.isdir(path):
                shutil.rmtree(path)
            elif os.path.isfile(path):
                os.remove(path)

    # 닫기 버튼 클릭 함수
    def close_window(self):
        self.master.destroy()  # Close the Tkinter window


class CheckDatabase:
    def __init__(self, master):
        self.connect_ini_file_name = 'connect.ini'

        self.master = master
        self.master.title("천마 탐지추적장치 장비 데이터 업로드  V1.4.0\n[2024-05-22 Released]")

        self.title = tk.Label(master, text="천마 탐지추적장치 장비 데이터 업로드  V1.4.0", font=("Arial", 15, "bold"))
        self.title.grid(row=0, column=0, sticky="nwe", pady=(20, 0), padx=(20, 20))

        self.close_button = ttk.Button(master, text="종료", command=self.close_window, width=5,
                                       style="CloseButton.TButton")
        self.close_button.grid(row=3, column=0, sticky='en', pady=(30, 0), padx=15)

        self.time_left = 10  # 시작 시간 설정 (초)

        self.label = tk.Label(master, text="")
        self.label.grid()

        self.update_label()

        style = ttk.Style()
        style.configure("CloseButton.TButton",
                        background="orange",  # 배경색
                        foreground="black",  # 텍스트 색상
                        font=("Helvetica", 10, "bold"),  # 폰트
                        padding=3,  # 안쪽 여백
                        relief="sunken",  # 외곽선 스타일
                        borderwidth=3,  # 외곽선 두께
                        )

        # 로딩 그래픽
        self.canvas_width = 15
        self.canvas_height = 15
        self.current_x = self.canvas_width // 2
        self.current_y = self.canvas_height // 2
        self.radius = 7
        self.angle = 0
        self.animation_canvas_a = tk.Canvas(master, width=self.canvas_width, height=self.canvas_height)
        self.animation_canvas_a.grid(row=2, column=0, pady=0, padx=80, sticky='w')
        self.animation_canvas_a.grid_remove()
        self.animation_canvas_b = tk.Canvas(master, width=self.canvas_width, height=self.canvas_height)
        self.animation_canvas_b.grid(row=3, column=0, pady=0, padx=80, sticky='w')
        self.animation_canvas_b.grid_remove()

        # Cheking 텍스트
        self.loading_label_a = tk.Label(master)
        self.loading_label_a.grid(row=2, column=0, pady=0, padx=120, sticky="w")
        self.loading_label_a.grid_remove()

        # Cheking 텍스트
        self.loading_label_b = tk.Label(master)
        self.loading_label_b.grid(row=3, column=0, pady=0, padx=120, sticky="w")
        self.loading_label_b.grid_remove()

        # 로딩중 출력
        self.dot = '.'
        self.load_label_a()
        self.load_label_b()
        self.draw_rotating_circle_a()
        self.draw_rotating_circle_b()

        # 성공
        self.success_a = tk.Label(master, text="Successfully connected", font=("Arial", 12, "bold"), fg="green")
        self.success_a.grid(row=2, column=0, sticky='w', pady=0, padx=120)
        self.success_b = tk.Label(master, text="Successfully connected", font=("Arial", 12, "bold"), fg="green")
        self.success_b.grid(row=3, column=0, sticky='w', pady=0, padx=120)
        self.success_a.grid_remove()
        self.success_b.grid_remove()

        # 실패
        self.fail_a = tk.Label(master, text="Connection failed", font=("Arial", 12, "bold"), fg="red")
        self.fail_a.grid(row=2, column=0, sticky='w', pady=0, padx=120)
        self.fail_b = tk.Label(master, text="Connection failed", font=("Arial", 12, "bold"), fg="red")
        self.fail_b.grid(row=3, column=0, sticky='w', pady=0, padx=120)
        self.fail_a.grid_remove()
        self.fail_b.grid_remove()

        # 텍스트
        self.information = tk.Label(master, text="현재 DB 접속 설정을 확인 중입니다. 기다려주십시요.", font=("Arial", 10))
        self.information.grid(row=1, column=0, sticky="nwe")

        # 텍스트
        self.waiting = tk.Label(master, text="프로그램이 시작됩니다.", font=("Arial", 10))
        self.waiting.grid(row=1, column=0, sticky="nwe")
        self.waiting.grid_remove()

        # 텍스트
        self.error = tk.Label(master, text="제대로 된 DB 정보를 입력해주십시요. 종료됩니다.", font=("Arial", 11, "bold"))
        self.error.grid(row=1, column=0, sticky="nwe")
        self.error.grid_remove()

        # 이미지 불러오기
        self.mysql_logo = "image/mysql.png"  # 이미지 파일 경로
        self.mysql_logo = tk.PhotoImage(file=self.mysql_logo).subsample(8)

        # 이미지를 표시할 라벨 생성
        self.mysql_logo_label = tk.Label(master, image=self.mysql_logo, borderwidth=1, relief="sunken")
        self.mysql_logo_label.config(width=40, height=40)
        self.mysql_logo_label.grid(row=2, column=0, sticky='w', pady=0, padx=20)

        # 이미지 불러오기
        self.influx_logo = "image/influx.png"  # 이미지 파일 경로
        self.influx_logo = tk.PhotoImage(file=self.influx_logo).subsample(8)

        # 이미지를 표시할 라벨 생성
        self.influx_logo_label = tk.Label(master, image=self.influx_logo, borderwidth=1, relief="sunken")
        self.influx_logo_label.config(width=40, height=40)
        self.influx_logo_label.grid(row=3, column=0, sticky='w', pady=0, padx=20)

        # 이미지 불러오기
        self.success_logo = "image/success.png"  # 이미지 파일 경로
        self.success_logo = tk.PhotoImage(file=self.success_logo).subsample(10)

        # 이미지를 표시할 라벨 생성
        self.success_logo_label = tk.Label(master, image=self.success_logo)
        self.success_logo_label.config(width=20, height=20)

        # 이미지를 표시할 라벨 생성
        self.success_logo_label2 = tk.Label(master, image=self.success_logo)
        self.success_logo_label2.config(width=20, height=20)
        # self.success_logo_label.grid(row=3, column=0, sticky='w', pady=0, padx=20)

        # 이미지 불러오기
        self.fail_logo = "image/fail.png"  # 이미지 파일 경로
        self.fail_logo = tk.PhotoImage(file=self.fail_logo).subsample(10)

        # 이미지를 표시할 라벨 생성
        self.fail_logo_label = tk.Label(master, image=self.fail_logo)
        self.fail_logo_label.config(width=20, height=20)
        # self.fail_logo.grid(row=3, column=0, sticky='w', pady=0, padx=20)

        # 연결 성공 여부 확인
        self.mysql_success_signal = False
        self.influx_success_signal = False

        self.mysql_connection = None
        self.influx_connection = None

        # DB 확인
        show_thread = threading.Thread(target=self.show)
        show_thread.start()
        check_thread = threading.Thread(target=self.check)
        check_thread.start()

    def close_window(self):
        roots.var = False
        roots.destroy()

    def func_dummy(self):
        print('dummy : processing...')

        # 연결 실패 시, 실패 코드 주고 연결 실패 알림 울리고 로그 남기고 인트로 창 종료

    def check(self):
        self.func_dummy()
        self.master.after(1000, self.mysql_check)
        self.influx_check()
        self.master.after(3000, self.info)
        self.master.after(7000, self.run_program)

    def run_program(self):
        if self.mysql_success_signal and self.influx_success_signal:
            roots.var = True
            roots.destroy()
        else:
            roots.var = False
            roots.destroy()

    def info(self):
        if self.mysql_success_signal and self.influx_success_signal:
            self.information.grid_remove()
            self.waiting.grid()

        else:
            self.information.grid_remove()
            self.error.grid()
            messagebox.showerror('연결 실패', 'log 폴더 내 오류를 확인해주십시요.')

    def show(self):
        self.animation_canvas_a.grid()
        self.animation_canvas_b.grid()
        self.loading_label_a.grid()
        self.loading_label_b.grid()

    def load_label_a(self):
        if self.dot == "........":
            self.dot = '.'
        self.dot += self.dot
        self.loading_label_a.config(text="Loading" + self.dot, font=("Helvetica", 12, "bold"))
        self.loader = self.loading_label_a.after(1000, self.load_label_a)

    def load_label_b(self):
        if self.dot == "........":
            self.dot = '.'
        self.dot += self.dot
        self.loading_label_b.config(text="Loading" + self.dot, font=("Helvetica", 12, "bold"))
        self.loader = self.loading_label_b.after(1000, self.load_label_b)

    def mysql_if_success(self):
        self.success_a.grid()
        self.loading_label_a.grid_remove()
        self.animation_canvas_a.grid_remove()
        self.mysql_success_signal = True
        self.animation_canvas_a.grid_remove()
        self.success_logo_label.grid(row=2, column=0, sticky='w', pady=0, padx=80)

    def influx_if_success(self):
        self.success_b.grid()
        self.loading_label_b.grid_remove()
        self.animation_canvas_b.grid_remove()
        self.influx_success_signal = True
        self.animation_canvas_b.grid_remove()
        self.success_logo_label2.grid(row=3, column=0, sticky='w', pady=0, padx=80)

    def mysql_if_failed(self):
        self.fail_a.grid()
        self.loading_label_a.grid_remove()
        self.animation_canvas_a.grid_remove()
        self.fail_logo_label.grid(row=2, column=0, sticky='w', pady=0, padx=80)

    def influx_if_failed(self):
        self.fail_b.grid()
        self.loading_label_b.grid_remove()
        self.animation_canvas_b.grid_remove()
        self.fail_logo_label.grid(row=3, column=0, sticky='w', pady=0, padx=80)

    def error_logging(self, e):
        err_msg = traceback.format_exc()
        log_file_path = f'log/{datetime.now().strftime("%Y년%m월%d일%H시%M분%S초")}ConnectionError.log'
        with open(log_file_path, 'w') as log_file:
            log_file.write(str(e))
            log_file.write('self.traceback_str')
            log_file.write(err_msg)

    def read_db_config(self, filename, section):
        parser = configparser.ConfigParser()
        parser.read(filename, 'UTF-8')
        db = {}
        if parser.has_section(section):
            items = parser.items(section)
            for item in items:
                db[item[0]] = item[1]
        else:
            raise Exception(f'{section} not found in the {filename} file')
        return db

    def mysql_check(self):
        try:
            mysql_config = self.read_db_config(self.connect_ini_file_name, 'MySQL')
            self.mysql_connection = mysql.connector.connect(
                host=mysql_config['host'],
                port=mysql_config['port'],
                user=mysql_config['user'],
                password=mysql_config['password'],
                database=mysql_config['database'],
                connect_timeout=15
            )
            self.mysql_if_success()
            self.mysql_success_signal = True
        except Exception as e:
            self.error_logging(e)
            self.mysql_if_failed()
            self.mysql_success_signal = False

    def influx_check(self):
        try:
            db_config = self.read_db_config()
            influx_exe_path = os.path.join(db_config['path'], 'influx.exe')
            if not os.path.exists(influx_exe_path):
                raise Exception(f"influx.exe not found in path: {db_config['path']}")
            self.influx_connection = InfluxDBClient(
                host=db_config['host'],
                port=db_config['port'],
                database=db_config['dbname'],
                timeout=15
            )

            self.influx_if_success()
            self.influx_success_signal = True
        except Exception as e:
            self.error_logging(e)
            self.influx_if_failed()
            self.influx_success_signal = False

    def draw_rotating_circle_a(self):
        self.animation_canvas_a.delete("rotating_circle")
        self.angle += 1
        x0 = self.current_x + self.radius * math.cos(math.radians(self.angle))
        y0 = self.current_y + self.radius * math.sin(math.radians(self.angle))
        x1 = self.current_x - self.radius * math.cos(math.radians(self.angle))
        y1 = self.current_y - self.radius * math.sin(math.radians(self.angle))
        self.animation_canvas_a.create_line(x0, y0, x1, y1, fill="orange", width=4, tags="rotating_circle")
        self.master.after(1, self.draw_rotating_circle_a)

    def draw_rotating_circle_b(self):
        self.animation_canvas_b.delete("rotating_circle")
        self.angle += 1
        x0 = self.current_x + self.radius * math.cos(math.radians(self.angle))
        y0 = self.current_y + self.radius * math.sin(math.radians(self.angle))
        x1 = self.current_x - self.radius * math.cos(math.radians(self.angle))
        y1 = self.current_y - self.radius * math.sin(math.radians(self.angle))
        self.animation_canvas_b.create_line(x0, y0, x1, y1, fill="orange", width=4, tags="rotating_circle")
        self.master.after(1, self.draw_rotating_circle_b)

    def update_label(self):
        if self.time_left > 0:
            self.time_left -= 1
            self.master.after(1000, self.update_label)  # 1초마다 update_label 함수 호출
        else:
            # 10초가 지나면 messagebox를 띄우고 애플리케이션을 종료합니다.
            roots.var = False
            messagebox.showerror("연결 실패 ( 지연 사유 )", "지연된 연결 상태로 인해, 프로그램이 종료됩니다.\n올바른 연결 정보를 입력해주십시요.")
            self.master.destroy()


""" CSV check """


class CheckCSV:
    def __init__(self, master):
        self.master = master
        self.master.title("천마 탐지추적장치 장비 데이터 업로드  V1.4.0\n[2024-05-22 Released]")

        self.title = tk.Label(master, text="천마 탐지추적장치 장비 데이터 업로드  V1.4.0", font=("Arial", 15, "bold"))
        self.title.grid(row=0, column=0, sticky="nwe", pady=(20, 0), padx=(20, 20))

        self.close_button = ttk.Button(master, text="종료", command=self.close_window, width=5,
                                       style="CloseButton.TButton")
        self.close_button.grid(row=2, column=0, sticky='en', pady=(30, 0), padx=15)

        self.time_left = 10  # 시작 시간 설정 (초)

        self.label = tk.Label(master, text="")
        self.label.grid()

        # self.update_label()

        style = ttk.Style()
        style.configure("CloseButton.TButton",
                        background="orange",  # 배경색
                        foreground="black",  # 텍스트 색상
                        font=("Helvetica", 10, "bold"),  # 폰트
                        padding=3,  # 안쪽 여백
                        relief="sunken",  # 외곽선 스타일
                        borderwidth=3,  # 외곽선 두께
                        )

        # 로딩 그래픽
        self.canvas_width = 25
        self.canvas_height = 25
        self.current_x = self.canvas_width // 2
        self.current_y = self.canvas_height // 2
        self.radius = 12.5
        self.angle = 0
        self.animation_canvas_a = tk.Canvas(master, width=self.canvas_width, height=self.canvas_height)
        self.animation_canvas_a.grid(row=2, column=0, pady=(15, 0), padx=(270, 0), sticky='n')
        # self.animation_canvas_a.grid_remove()

        # 텍스트
        self.information = tk.Label(master, text="CSV 변환/결합 프로그램이 시작됩니다.", font=("Arial", 10))
        self.information.grid(row=2, column=0, sticky="n", pady=(20, 0))

        # 이미지 불러오기
        self.excel_logo = "image/excel.png"  # 이미지 파일 경로
        self.excel_logo = tk.PhotoImage(file=self.excel_logo).subsample(8)

        # 이미지를 표시할 라벨 생성
        self.mysql_logo_label = tk.Label(master, image=self.excel_logo, borderwidth=1, relief="sunken")
        self.mysql_logo_label.config(width=50, height=50)
        self.mysql_logo_label.grid(row=2, column=0, sticky='w', pady=0, padx=20)

        self.draw_rotating_circle_a()

        thread1 = threading.Thread(target=self.run)
        thread1.start()

    def run(self):
        self.master.after(2000, self.run_program)

    def run_program(self):
        croots.var = True
        croots.destroy()

    def draw_rotating_circle_a(self):
        self.animation_canvas_a.delete("rotating_circle")
        self.angle += 1
        x0 = self.current_x + self.radius * math.cos(math.radians(self.angle))
        y0 = self.current_y + self.radius * math.sin(math.radians(self.angle))
        x1 = self.current_x - self.radius * math.cos(math.radians(self.angle))
        y1 = self.current_y - self.radius * math.sin(math.radians(self.angle))
        self.animation_canvas_a.create_line(x0, y0, x1, y1, fill="orange", width=4, tags="rotating_circle")
        self.master.after(1, self.draw_rotating_circle_a)

    def close_window(self):
        croots.var = False
        croots.destroy()


def 로딩화면설정():
    roots.overrideredirect(True)
    window_widths = roots.winfo_reqwidth() * 25 // 11
    window_heights = roots.winfo_reqheight() * 9 // 10
    screen_widths = roots.winfo_screenwidth()
    screen_heights = roots.winfo_screenheight()
    x_positions = (screen_widths - window_widths) // 2
    y_positions = (screen_heights - window_heights) // 2
    roots.geometry(f"{window_widths}x{window_heights}+{x_positions}+{y_positions}")
    roots.update()
    roots.attributes("-topmost", True)
    roots.mainloop()
    return roots.var


def CSV_로딩화면설정():
    croots.overrideredirect(True)
    window_widths = croots.winfo_reqwidth() * 25 // 11
    window_heights = croots.winfo_reqheight() * 6 // 10
    screen_widths = croots.winfo_screenwidth()
    screen_heights = croots.winfo_screenheight()
    x_positions = (screen_widths - window_widths) // 2
    y_positions = (screen_heights - window_heights) // 2
    croots.geometry(f"{window_widths}x{window_heights}+{x_positions}+{y_positions}")
    croots.update()
    croots.attributes("-topmost", True)
    croots.mainloop()
    return croots.var


if __name__ == '__main__':
    config = configparser.ConfigParser()
    config.read('connect.ini', 'UTF-8')
    get_type = config.get('Program', 'type')
    if get_type == 'ALL':
        print('dummy: this type is ALL type')
        try:
            roots = tk.Tk()
            apps = CheckDatabase(roots)
            roots.var = True
            if 로딩화면설정():
                root = tk.Tk()
                app = FileListApp(root)
                print('dummy: can run second program')
                window_width = root.winfo_reqwidth() * 44 // 10
                window_height = root.winfo_reqheight() * 28 // 10
                screen_width = root.winfo_screenwidth()
                screen_height = root.winfo_screenheight()
                x_position = (screen_width - window_width) // 2
                y_position = (screen_height - window_height) // 3
                root.geometry(f"{window_width}x{window_height}+{x_position}+{y_position}")
                root.update()
                root.mainloop()
            else:
                print('dummy: cant run second program')
        except Exception as E:
            if 'while executing' in str(E):
                pass
            else:
                err_msg = traceback.format_exc()
                log_file_path = f'log/{datetime.now().strftime("%Y년%m월%d일%H시%M분%S초")}ProgramError.log'
                with open(log_file_path, 'w') as log_file:
                    log_file.write(str(E))
                    log_file.write('self.traceback_str')
                    log_file.write(err_msg)

    elif get_type == 'CSV':
        print('dummy: this type is CSV type')
        try:
            croots = tk.Tk()
            apps = CheckCSV(croots)
            croots.var = True
            if CSV_로딩화면설정():
                root = tk.Tk()
                app = FileListApp(root)
                print('dummy: can run second program')
                window_width = root.winfo_reqwidth() * 44 // 10
                window_height = root.winfo_reqheight() * 28 // 10
                screen_width = root.winfo_screenwidth()
                screen_height = root.winfo_screenheight()
                x_position = (screen_width - window_width) // 2
                y_position = (screen_height - window_height) // 3
                root.geometry(f"{window_width}x{window_height}+{x_position}+{y_position}")
                root.update()
                root.mainloop()
            else:
                print('dummy: cant run second program')
        except Exception as E:
            if 'while executing' in str(E):
                pass
            else:
                err_msg = traceback.format_exc()
                log_file_path = f'log/{datetime.now().strftime("%Y년%m월%d일%H시%M분%S초")}ProgramError.log'
                with open(log_file_path, 'w') as log_file:
                    log_file.write(str(E))
                    log_file.write('self.traceback_str')
                    log_file.write(err_msg)
